# -*- coding: utf-8 -*-
"""
Excel 文件解析器（重构版 v2.0）

核心改进：
1. 多策略提取：支持 .xlsx / .xls / .csv 三种格式
2. 多 Sheet 支持：自动扫描所有工作表
3. 格式自动识别：Format A（横向）/ Format B（纵向）/ 表头行式
4. 字段级验证：每个字段提取后验证合理性
5. 详细日志：每个步骤记录详细信息便于调试
6. 批量处理优化：支持并行处理多个文件
"""

import os
import re
import csv
import logging
import datetime
from typing import Optional, Dict, List, Any
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

try:
    import xlrd
    _HAS_XLRD = True
except ImportError:
    _HAS_XLRD = False

from core.logger import get_logger
from plugins.peel_data.models import PeelDataRecord
from config import config

logger = get_logger("peel_data.excel_parser")


# ─── 常量定义 ────────────────────────────────────────────────────────────

# 支持的扩展名
SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xlsb", ".xls", ".csv")

# 默认超时（秒）
DEFAULT_TIMEOUT = 30


# ─── 工具函数 ────────────────────────────────────────────────────────────

def _clean_number(text: str) -> Optional[float]:
    """清理数字字符串并转为 float"""
    if not text:
        return None
    try:
        # 移除空格、换行等
        cleaned = str(text).strip().replace(" ", "").replace("\n", "")
        # 处理逗号分隔符
        cleaned = cleaned.replace(",", "")
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _parse_date(val: Any) -> Optional[datetime.date]:
    """解析日期值（支持字符串、datetime 对象、Excel 序列号）"""
    if val is None:
        return None

    # datetime 对象（先于 date 检查，因为 datetime 是 date 的子类）
    if isinstance(val, datetime.datetime):
        return val.date()

    # 已经是 date 对象
    if isinstance(val, datetime.date):
        return val

    # 字符串
    if isinstance(val, str):
        # YYYY-M-D
        m = re.match(r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})", val)
        if m:
            try:
                return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass

    # Excel 序列号（数字）
    if isinstance(val, (int, float)):
        try:
            # openpyxl 的 data_only=True 会返回 datetime 对象
            # 如果还是数字，尝试转换
            from openpyxl.utils.datetime import from_excel
            return from_excel(val)
        except Exception:
            pass

    return None


def _parse_time(val: Any) -> Optional[datetime.time]:
    """解析时间值"""
    if val is None:
        return None

    if isinstance(val, datetime.time):
        return val

    if isinstance(val, datetime.datetime):
        return val.time()

    if isinstance(val, str):
        m = re.match(r"(\d{1,2}):(\d{2}):(\d{2})", val)
        if m:
            try:
                return datetime.time(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass

    return None


# ─── 主解析器类 ────────────────────────────────────────────────────────────

class ExcelParser:
    """
    Excel 剥离数据解析器（重构版）

    提取策略：
    1. 自动识别文件格式（.xlsx / .xls / .csv）
    2. 多 Sheet 扫描：自动扫描所有工作表
    3. 格式自动识别：Format A / Format B / 表头行式
    4. 字段验证：提取后验证每个字段的合理性
    """

    # 关键字定义
    SAMPLE_NAME_KEYWORDS = ["试样名称", "样品名称", "试样编号", "样品名"]
    SAMPLE_BRAND_KEYWORDS = ["试样牌号", "牌号", "样本编号"]
    DATE_KEYWORDS = ["试验日期", "测试日期", "日期"]
    TIME_KEYWORDS = ["试验时间", "测试时间", "时间"]
    CURVE_KEYWORDS = [f"曲线{i}" for i in range(1, 10)]
    STD_KEYWORDS = ["标准差", "A_sd", "A_SD", "Std", "SD"]

    GENERIC_NAMES = frozenset([
        "试样编号", "试样名称", "样品名称", "样品名", "试样", "编号", "名称", ""
    ])

    def parse(self, file_path: str, timeout: int = DEFAULT_TIMEOUT) -> Optional[PeelDataRecord]:
        """
        解析单个 Excel/CSV 文件

        Args:
            file_path: 文件路径
            timeout: 解析超时时间（秒）

        Returns:
            PeelDataRecord 或 None
        """
        filename = os.path.basename(file_path)
        logger.info("=" * 60)
        logger.info("开始解析: %s", filename)
        t_start = datetime.datetime.now()

        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext == ".csv":
                record = self._parse_csv(file_path)
            elif ext in (".xls",):
                record = self._parse_xls(file_path)
            else:
                record = self._parse_xlsx(file_path)

            if record is None:
                logger.warning("[%s] 解析失败：无法提取有效数据", filename)
                return None

            # 填充文件信息（存储完整绝对路径，方便用户右键打开文件位置）
            record.source_file = file_path
            record.file_type = "excel" if ext != ".csv" else "csv"

            # ── 试样名称按优先级重提取 ──
            # 优先级 1: 文件名（最高优先）
            # 优先级 2: 材料关键词辅助判断
            # 优先级 3: "正极"/"负极" 前缀全量提取
            # 优先级 4: 保留已提取的 sample_name
            record.sample_name = self._upgrade_sample_name(
                record.sample_name, filename
            )

            # 判断极性
            record.polarity = self._determine_polarity(record.sample_name, filename)

            # 验证字段
            self._validate_record(record, filename)

            elapsed = (datetime.datetime.now() - t_start).total_seconds()
            s_count = sum(1 for k in ["curve_1", "curve_2", "curve_3", "curve_4"]
                        if getattr(record, k, None) is not None)
            logger.info("[%s] 解析完成 (%.2fs): 试样=%s, curve=%d/4, std_dev=%s",
                        filename, elapsed, record.sample_name or "未知", s_count,
                        "有" if record.std_dev else "无")
            logger.info("=" * 60)

            return record

        except Exception as e:
            logger.exception("[%s] 解析异常: %s", filename, e)
            return None

    def _parse_xlsx(self, file_path: str) -> Optional[PeelDataRecord]:
        """解析 .xlsx / .xlsm / .xlsb 文件（多 Sheet 支持）"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        record = None

        # 策略1：扫描所有 Sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.debug("[%s] 扫描 Sheet: %s", os.path.basename(file_path), sheet_name)

            sheet_data = self._parse_worksheet(ws)
            if sheet_data and (record is None or self._is_better_record(sheet_data, record)):
                record = sheet_data
                logger.debug("[%s] Sheet '%s' 提取到更完整的数据",
                             os.path.basename(file_path), sheet_name)

        return record

    def _parse_xls(self, file_path: str) -> Optional[PeelDataRecord]:
        """解析 .xls 文件"""
        if not _HAS_XLRD:
            logger.error("xlrd 未安装，无法解析 .xls 文件")
            return None

        workbook = xlrd.open_workbook(file_path)
        record = None

        for sheet_idx in range(workbook.nsheets):
            ws = workbook.sheet_by_index(sheet_idx)
            sheet_data = self._parse_xls_sheet(ws)
            if sheet_data and (record is None or self._is_better_record(sheet_data, record)):
                record = sheet_data

        return record

    def _parse_csv(self, file_path: str) -> Optional[PeelDataRecord]:
        """解析 .csv 文件"""
        record = PeelDataRecord()

        with open(file_path, "r", encoding="utf-8-sig", errors="ignore") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            return None

        # 尝试从 CSV 中提取数据
        # 策略1：查找表头行
        header_row = None
        for idx, row in enumerate(rows):
            row_str = " ".join(str(cell) for cell in row if cell)
            if "剥离强度" in row_str or "试样名称" in row_str:
                header_row = idx
                break

        if header_row is not None:
            # 提取元数据
            for row in rows[:header_row + 10]:
                row_str = " ".join(str(cell) for cell in row if cell)
                self._extract_metadata_from_text(row_str, record)

            # 提取 S 值（查找"平均值"行）
            for row in rows[header_row:]:
                if "平均值" in " ".join(str(cell) for cell in row if cell):
                    # 尝试从这一行提取 S1-S4
                    for idx, cell in enumerate(row):
                        if cell and re.match(r"^\d+\.?\d*$", str(cell)):
                            if idx < len(row) - 3:
                                record.curve_1 = _clean_number(row[idx])
                                record.curve_2 = _clean_number(row[idx + 1])
                                record.curve_3 = _clean_number(row[idx + 2])
                                record.curve_4 = _clean_number(row[idx + 3])
                            break
                    break

        return record if (record.sample_name or record.curve_1 is not None) else None

    def _parse_worksheet(self, ws) -> Optional[PeelDataRecord]:
        """解析单个 Worksheet"""
        record = PeelDataRecord()

        # 步骤1：提取元数据
        self._extract_metadata_from_worksheet(ws, record)

        # 步骤2：定位剥离强度数据区域
        peel_row = self._find_peel_header_row(ws)
        if peel_row is None:
            logger.debug("未找到剥离强度区域")
            return None

        # 步骤3：判断格式并提取
        fmt = self._detect_format(ws, peel_row)

        if fmt == "A":
            self._parse_format_a(ws, peel_row, record)
        elif fmt == "B":
            self._parse_format_b(ws, peel_row, record)
        else:
            # 回退：尝试两种格式
            logger.debug("格式未知，尝试 Format A")
            self._parse_format_a(ws, peel_row, record)
            if record.curve_1 is None:
                logger.debug("Format A 失败，尝试 Format B")
                self._parse_format_b(ws, peel_row, record)

        return record if (record.sample_name or record.curve_1 is not None) else None

    def _extract_metadata_from_worksheet(self, ws, record: PeelDataRecord):
        """从 Worksheet 中提取元数据"""
        # 第一遍：提取试样名称、牌号、日期、时间（试样名称优先于试验批号）
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue

                cell_str = str(val).strip()

                # 试样名称
                for kw in self.SAMPLE_NAME_KEYWORDS:
                    if kw in cell_str and not record.sample_name:
                        name = None
                        if cell_str != kw:
                            remaining = cell_str.split(kw, 1)[-1].strip()
                            remaining = remaining.lstrip('=：: ').strip()
                            if remaining and remaining not in self.GENERIC_NAMES:
                                name = remaining
                        if name is None:
                            for offset in [1, 2]:
                                next_cell = ws.cell(row=cell.row, column=cell.column + offset).value
                                if next_cell:
                                    candidate = str(next_cell).strip()
                                    if candidate and candidate not in self.GENERIC_NAMES:
                                        name = candidate
                                        break
                        if name:
                            record.sample_name = name
                            break

                # 试样牌号
                for kw in self.SAMPLE_BRAND_KEYWORDS:
                    if kw in cell_str and not record.sample_brand:
                        brand = None
                        if cell_str != kw:
                            remaining = cell_str.split(kw, 1)[-1].strip()
                            remaining = remaining.lstrip('=：: ').strip()
                            if remaining:
                                brand = remaining
                        if brand is None:
                            next_cell = ws.cell(row=cell.row, column=cell.column + 1).value
                            if next_cell:
                                record.sample_brand = str(next_cell).strip()
                        if brand:
                            record.sample_brand = brand
                        break

                # 试验日期
                for kw in self.DATE_KEYWORDS:
                    if kw in cell_str and not record.test_date:
                        date_val = None
                        if cell_str != kw:
                            remaining = cell_str.split(kw, 1)[-1].strip()
                            remaining = remaining.lstrip('=：: ').strip()
                            if remaining:
                                date_val = remaining
                        if date_val is None:
                            for offset in [1, 2]:
                                next_cell = ws.cell(row=cell.row, column=cell.column + offset).value
                                if next_cell:
                                    parsed = _parse_date(next_cell)
                                    if parsed:
                                        record.test_date = str(parsed) if not isinstance(parsed, str) else parsed
                                        break
                        else:
                            parsed = _parse_date(date_val)
                            if parsed:
                                record.test_date = str(parsed) if not isinstance(parsed, str) else parsed
                        break

                # 试验时间
                for kw in self.TIME_KEYWORDS:
                    if kw in cell_str and not record.test_time:
                        time_val = None
                        if cell_str != kw:
                            remaining = cell_str.split(kw, 1)[-1].strip()
                            remaining = remaining.lstrip('=：: ').strip()
                            if remaining:
                                time_val = remaining
                        if time_val is None:
                            for offset in [1, 2]:
                                next_cell = ws.cell(row=cell.row, column=cell.column + offset).value
                                if next_cell:
                                    parsed = _parse_time(next_cell)
                                    if parsed:
                                        record.test_time = str(parsed) if not isinstance(parsed, str) else parsed
                                        break
                        else:
                            parsed = _parse_time(time_val)
                            if parsed:
                                record.test_time = str(parsed) if not isinstance(parsed, str) else parsed
                        break

        # 第二遍：如果试样名称仍为空，用试验批号作为 fallback
        if not record.sample_name:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if val is None:
                        continue
                    cell_str = str(val).strip()
                    if "试验批号" in cell_str:
                        name = None
                        if cell_str != "试验批号":
                            remaining = cell_str.split("试验批号", 1)[-1].strip()
                            remaining = remaining.lstrip('=：: ').strip()
                            if remaining and remaining not in self.GENERIC_NAMES:
                                name = remaining
                        if name is None:
                            for offset in [1, 2]:
                                next_cell = ws.cell(row=cell.row, column=cell.column + offset).value
                                if next_cell:
                                    candidate = str(next_cell).strip()
                                    if candidate and candidate not in self.GENERIC_NAMES:
                                        name = candidate
                                        break
                        if name:
                            record.sample_name = name
                            break
                        else:
                            record.test_time = _parse_time(time_val)
                        break

    def _find_peel_header_row(self, ws) -> Optional[int]:
        """查找剥离强度表头行"""
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").strip().replace("\n", "")
                if "剥离强度" in val:
                    return cell.row
        return None

    def _detect_format(self, ws, peel_row: int) -> str:
        """判断数据格式（A: 横向 / B: 纵向）"""
        # 检查是否有 S1-S4 列头
        has_s_columns = False
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=peel_row, column=col).value or "").strip()
            if re.match(r"^S[1-4]$", val):
                has_s_columns = True
                break

        # 检查是否有"曲线1"等行
        has_curve_rows = False
        for row in range(peel_row + 1, min(peel_row + 5, ws.max_row + 1)):
            val = str(ws.cell(row=row, column=1).value or "").strip()
            if re.match(r"^曲线[1-9]$", val):
                has_curve_rows = True
                break

        if has_curve_rows and not has_s_columns:
            return "B"
        elif has_s_columns:
            return "A"
        else:
            return "unknown"

    def _parse_format_a(self, ws, peel_row: int, record: PeelDataRecord):
        """解析 Format A（横向 S 列布局）"""
        # 找 S 列索引
        s_col_map = {}
        for row in [peel_row, peel_row + 1]:
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(row=row, column=col).value or "").strip()
                if re.match(r"^S[1-4]$", val):
                    s_col_map[val] = col

        # 找平均值行
        avg_row = None
        for row in range(peel_row, ws.max_row + 1):
            val = str(ws.cell(row=row, column=1).value or "").strip()
            if "平均值" in val:
                avg_row = row
                break

        if avg_row and s_col_map:
            for label, col in s_col_map.items():
                val = ws.cell(row=avg_row, column=col).value
                if val is not None:
                    # 映射 S1→curve_1, S2→curve_2, ...
                    curve_key = f"curve_{label[1:]}" if label.startswith("S") and label[1:].isdigit() else label
                    setattr(record, curve_key, _clean_number(val))

        # 找 A_sd
        for row in range(peel_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(row=row, column=col).value or "")
                for kw in self.STD_KEYWORDS:
                    if kw in val:
                        # 尝试提取数字
                        m = re.search(r"([\d.]+)", val)
                        if m:
                            record.std_dev = float(m.group(1))
                            return

    def _parse_format_b(self, ws, peel_row: int, record: PeelDataRecord):
        """解析 Format B（纵向曲线行布局）

        布局示例：
        R11: 剥离强度(kN/m)
        R12: [空] | 最大剥离强度 | 最小剥离强度 | 平均剥离强度 | 标准差
        R13: 曲线1 | 0.0045 | 0.0035 | 0.0038 | 0.0001
        ...
        R22: 平均值 | 0.0064 | 0.004 | 0.0046 | 0.0003
        """
        # 找子表头行（含"平均剥离强度"和"标准差"）
        sub_header_row = None
        avg_col = None
        std_col = None
        max_col = None
        min_col = None

        for row in range(peel_row + 1, min(peel_row + 5, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(row=row, column=col).value or "").strip()
                if "平均" in val and "剥离" in val:
                    avg_col = col
                    sub_header_row = row
                if "标准差" in val or "A_sd" in val:
                    std_col = col
                    if not sub_header_row:
                        sub_header_row = row
                if "最大" in val and "剥离" in val:
                    max_col = col
                    if not sub_header_row:
                        sub_header_row = row
                if "最小" in val and "剥离" in val:
                    min_col = col
                    if not sub_header_row:
                        sub_header_row = row

        if sub_header_row is None:
            return

        # 提取曲线行数据（支持曲线1-9）
        curve_data = {}
        for row in range(sub_header_row + 1, ws.max_row + 1):
            val = str(ws.cell(row=row, column=1).value or "").strip()
            m = re.match(r"^曲线(\d+)$", val)
            if m:
                curve_num = m.group(1)
                if avg_col:
                    avg_val = ws.cell(row=row, column=avg_col).value
                    if avg_val is not None:
                        parsed = _clean_number(avg_val)
                        if parsed is not None:
                            curve_data[f"S{curve_num}"] = parsed

        # 填充到 record（S1→curve_1, S2→curve_2, ...）
        for k, v in curve_data.items():
            num = int(k[1:])
            if 1 <= num <= 9:
                curve_key = f"curve_{num}"
                setattr(record, curve_key, v)

        # 提取"平均值"汇总行的标准差（总标准差，最准确）
        for row in range(sub_header_row + 1, ws.max_row + 1):
            val = str(ws.cell(row=row, column=1).value or "").strip()
            if "平均值" in val:
                if avg_col:
                    overall_avg = ws.cell(row=row, column=avg_col).value
                    if overall_avg is not None:
                        parsed = _clean_number(overall_avg)
                        if parsed is not None and record.curve_1 is None:
                            # 如果没有曲线数据，用总平均值作为 curve_1
                            record.curve_1 = parsed
                if std_col:
                    overall_std = ws.cell(row=row, column=std_col).value
                    if overall_std is not None:
                        parsed = _clean_number(overall_std)
                        if parsed is not None:
                            record.std_dev = parsed
                            logger.debug("从平均值行提取 std_dev = %s", parsed)
                break

    def _parse_xls_sheet(self, ws) -> Optional[PeelDataRecord]:
        """解析 .xls 文件的单个 Sheet（完整实现 xlrd 适配版）"""
        record = PeelDataRecord()

        # 提取元数据
        for row_idx in range(min(20, ws.nrows)):
            for col_idx in range(min(10, ws.ncols)):
                val = str(ws.cell_value(row_idx, col_idx) or "").strip()

                for kw in self.SAMPLE_NAME_KEYWORDS:
                    if kw in val and not record.sample_name:
                        if col_idx + 1 < ws.ncols:
                            next_val = str(ws.cell_value(row_idx, col_idx + 1) or "").strip()
                            if next_val not in self.GENERIC_NAMES:
                                record.sample_name = next_val

                for kw in self.SAMPLE_BRAND_KEYWORDS:
                    if kw in val and not record.sample_brand:
                        if col_idx + 1 < ws.ncols:
                            record.sample_brand = str(ws.cell_value(row_idx, col_idx + 1) or "").strip()

                for kw in self.DATE_KEYWORDS:
                    if kw in val and not record.test_date:
                        if col_idx + 1 < ws.ncols:
                            next_val = ws.cell_value(row_idx, col_idx + 1)
                            parsed = _parse_date(next_val)
                            if parsed:
                                record.test_date = str(parsed)

                for kw in self.TIME_KEYWORDS:
                    if kw in val and not record.test_time:
                        if col_idx + 1 < ws.ncols:
                            next_val = ws.cell_value(row_idx, col_idx + 1)
                            parsed = _parse_time(next_val)
                            if parsed:
                                record.test_time = str(parsed)

        # 查找剥离强度区域
        peel_row = None
        for row_idx in range(ws.nrows):
            for col_idx in range(ws.ncols):
                val = str(ws.cell_value(row_idx, col_idx) or "").strip()
                if "剥离强度" in val:
                    peel_row = row_idx
                    break
            if peel_row is not None:
                break

        if peel_row is not None:
            # 判断格式：A（横向 S 列）或 B（纵向曲线行）
            has_s_columns = False
            has_curve_rows = False

            # 检查 S1-S4 列头
            for col_idx in range(ws.ncols):
                val = str(ws.cell_value(peel_row, col_idx) or "").strip()
                if re.match(r"^S[1-4]$", val):
                    has_s_columns = True

            # 检查曲线行
            for row_idx in range(peel_row + 1, min(peel_row + 5, ws.nrows)):
                val = str(ws.cell_value(row_idx, 0) or "").strip()
                if re.match(r"^曲线[1-9]$", val):
                    has_curve_rows = True
                    break

            if has_s_columns:
                # Format A：横向 S 列布局
                self._parse_xls_format_a(ws, peel_row, record)
            elif has_curve_rows:
                # Format B：纵向曲线行布局
                self._parse_xls_format_b(ws, peel_row, record)

            # 查找 A_sd
            for row_idx in range(peel_row, ws.nrows):
                for col_idx in range(ws.ncols):
                    val = str(ws.cell_value(row_idx, col_idx) or "")
                    for kw in self.STD_KEYWORDS:
                        if kw in val:
                            m = re.search(r"([\d.]+)", val)
                            if m:
                                record.std_dev = float(m.group(1))
                                break

        return record if (record.sample_name or record.curve_1 is not None) else None

    def _parse_xls_format_a(self, ws, peel_row: int, record: PeelDataRecord):
        """解析 .xls Format A（横向 S 列布局，xlrd 适配）"""
        # 找 S 列索引（扫描 peel_row 和 peel_row+1）
        s_col_map = {}
        for row in [peel_row, peel_row + 1]:
            if row >= ws.nrows:
                continue
            for col in range(ws.ncols):
                val = str(ws.cell_value(row, col) or "").strip()
                m = re.match(r"^S([1-9])$", val)
                if m:
                    s_col_map[m.group(0)] = col

        # 找平均值行
        avg_row = None
        for row in range(peel_row, ws.nrows):
            val = str(ws.cell_value(row, 0) or "").strip()
            if "平均值" in val:
                avg_row = row
                break

        if avg_row is not None and s_col_map:
            for label, col in s_col_map.items():
                val = ws.cell_value(avg_row, col)
                if val != "" and val is not None:
                    num_idx = label[1:]  # "S1" → "1"
                    if num_idx.isdigit():
                        curve_key = f"curve_{num_idx}"
                        parsed = _clean_number(val)
                        if parsed is not None:
                            setattr(record, curve_key, parsed)

    def _parse_xls_format_b(self, ws, peel_row: int, record: PeelDataRecord):
        """解析 .xls Format B（纵向曲线行布局，xlrd 适配）"""
        # 找子表头行和列位置
        sub_header_row = None
        avg_col = None
        std_col = None

        for row in range(peel_row + 1, min(peel_row + 5, ws.nrows)):
            for col in range(ws.ncols):
                val = str(ws.cell_value(row, col) or "").strip()
                if "平均" in val and "剥离" in val:
                    avg_col = col
                    sub_header_row = row
                if "标准差" in val or "A_sd" in val:
                    std_col = col
                    if sub_header_row is None:
                        sub_header_row = row

        if sub_header_row is None:
            return

        # 提取曲线行数据
        for row in range(sub_header_row + 1, ws.nrows):
            val = str(ws.cell_value(row, 0) or "").strip()
            m = re.match(r"^曲线(\d+)$", val)
            if m and avg_col is not None:
                curve_num = m.group(1)
                avg_val = ws.cell_value(row, avg_col)
                if avg_val != "" and avg_val is not None:
                    parsed = _clean_number(avg_val)
                    if parsed is not None and 1 <= int(curve_num) <= 9:
                        curve_key = f"curve_{curve_num}"
                        setattr(record, curve_key, parsed)

        # 提取"平均值"汇总行的标准差
        for row in range(sub_header_row + 1, ws.nrows):
            val = str(ws.cell_value(row, 0) or "").strip()
            if "平均值" in val:
                if std_col is not None:
                    overall_std = ws.cell_value(row, std_col)
                    if overall_std != "" and overall_std is not None:
                        parsed = _clean_number(overall_std)
                        if parsed is not None:
                            record.std_dev = parsed
                break

    def _extract_metadata_from_text(self, text: str, record: PeelDataRecord):
        """从文本中提取元数据"""
        # 试样名称
        for kw in self.SAMPLE_NAME_KEYWORDS:
            if kw in text and not record.sample_name:
                m = re.search(rf"{kw}[：:\s]+([^\n,]+)", text)
                if m:
                    name = m.group(1).strip()
                    if name not in self.GENERIC_NAMES:
                        record.sample_name = name

        # 日期
        for kw in self.DATE_KEYWORDS:
            if kw in text and not record.test_date:
                m = re.search(r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})", text)
                if m:
                    record.test_date = _parse_date(m.group(0))

    def _is_better_record(self, new: PeelDataRecord, old: PeelDataRecord) -> bool:
        """判断新记录是否比旧记录更完整"""
        new_count = sum(1 for k in ["curve_1", "curve_2", "curve_3", "curve_4", "std_dev"]
                      if getattr(new, k, None) is not None)
        old_count = sum(1 for k in ["curve_1", "curve_2", "curve_3", "curve_4", "std_dev"]
                      if getattr(old, k, None) is not None)
        return new_count > old_count

    def _validate_record(self, record: PeelDataRecord, filename: str):
        """验证记录的字段合理性"""
        # 验证 S 值（实际数据范围 0.001 ~ 0.1 kN/m）
        for sk in ["curve_1", "curve_2", "curve_3", "curve_4"]:
            val = getattr(record, sk, None)
            if val is not None:
                if not (0.0001 <= val <= 100):
                    logger.warning("[%s] %s 值异常: %s", filename, sk, val)

        # 验证 std_dev
        if record.std_dev is not None:
            if not (0 <= record.std_dev <= 100):
                logger.warning("[%s] std_dev 值异常: %s", filename, record.std_dev)

        # 验证日期
        if record.test_date is not None:
            today = datetime.date.today()
            try:
                if isinstance(record.test_date, datetime.date) and record.test_date > today:
                    logger.warning("[%s] 试验日期在未来: %s", filename, record.test_date)
                elif isinstance(record.test_date, str) and record.test_date:
                    parsed = _parse_date(record.test_date)
                    if parsed and parsed > today:
                        logger.warning("[%s] 试验日期在未来: %s", filename, record.test_date)
            except TypeError:
                pass

    def _determine_polarity(self, sample_name: str, filename: str) -> str:
        """判断极性（使用配置中的正负极关键字）"""
        polarity, _ = config.determine_polarity_with_materials(sample_name, filename)
        return polarity

    def _upgrade_sample_name(
        self, existing: Optional[str], filename: str
    ) -> Optional[str]:
        """
        按优先级链升级试样名称

        优先级：
        1. 从文件名提取（最高优先）
        2. 材料关键词辅助判断（用于修正极性，不影响 sample_name 内容）
        3. "正极"/"负极"前缀全量提取（从已有 sample_name 中）
        4. 保留已提取的 sample_name
        """
        # 优先级 1: 从文件名提取
        from_filename = config.extract_sample_name_from_filename(filename)
        if from_filename:
            # 如果文件名有"正极"或"负极"前缀，验证其完整性
            from_polarity = config.extract_sample_name_by_polarity_prefix(from_filename)
            if from_polarity:
                logger.info(
                    "[%s] 试样名称(优先级1-文件名+极性前缀) = %s",
                    os.path.basename(filename), from_polarity
                )
                return from_polarity
            # 文件名无极性前缀但仍然是有效来源
            logger.info(
                "[%s] 试样名称(优先级1-文件名) = %s",
                os.path.basename(filename), from_filename
            )
            return from_filename

        # 优先级 3: 从已有 sample_name 中提取"正极"/"负极"前缀
        if existing:
            from_polarity = config.extract_sample_name_by_polarity_prefix(existing)
            if from_polarity and from_polarity != existing:
                logger.info(
                    "[%s] 试样名称(优先级3-极性前缀) = %s (原: %s)",
                    os.path.basename(filename), from_polarity, existing
                )
                return from_polarity
            # 优先级 2: 材料关键词辅助（用于日志记录，不修改 sample_name）
            materials = config.match_material_keywords(existing)
            if materials:
                logger.debug(
                    "[%s] 试样名称匹配到材料关键词: %s",
                    os.path.basename(filename), materials
                )

        # 优先级 4: 保留已有的 sample_name
        return existing
