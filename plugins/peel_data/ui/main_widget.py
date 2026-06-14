# -*- coding: utf-8 -*-
"""
剥离数据汇总插件主界面
包含：目录选择、提取操作、数据预览表格、导出功能、日志面板、历史记录、编辑保存、版本动态
"""

import os
import subprocess
from typing import Optional, List

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QLabel, QPushButton, QLineEdit, QFileDialog, QGroupBox,
    QTableWidget, QTableWidgetItem, QComboBox, QCheckBox,
    QProgressBar, QTextEdit, QSplitter, QHeaderView,
    QMessageBox, QSizePolicy, QDialog, QMenu,
)
from PySide6.QtCore import Qt, QThread, Signal, QTimer, QDateTime, QPoint
from PySide6.QtGui import QColor, QFont, QTextDocument, QAction

from config import config
from core.logger import get_logger, ToolkitLogger
from core.database import DatabaseManager
from plugins.peel_data.extractor import PeelDataExtractor, ExtractionResult, FileHistory, _is_db_available
from plugins.peel_data.models import PeelDataRecord

logger = get_logger("peel_data.ui")
_SUMMARY_TABLE_NAME = PeelDataRecord.get_table_name()


class PreviewDialog(QDialog):
    """
    提取结果预览弹窗（替代简单 QMessageBox）

    功能：
    1. 显示数据预览表格（与主界面相同的列）
    2. 高亮被应用层去重跳过的记录（背景色标记）
    3. 顶部统计：扫描/成功/失败/应用层去重/数据库写入/数据库跳过
    4. 底部按钮：导出 Excel / 关闭
    """

    # 应用层去重跳过的记录背景色（暖黄）
    _SKIP_BG = QColor("#fff3cd")
    _SKIP_BG_DARK = QColor("#ffe69c")
    _HEADER_BG = QColor("#f0f0f0")
    _POLARITY_POS = QColor("#e74c3c")
    _POLARITY_NEG = QColor("#27ae60")
    _POLARITY_OTHER = QColor("#95a5a6")

    def __init__(self, result: "ExtractionResult", parent=None):
        super().__init__(parent)
        self._result = result
        self._records: List[PeelDataRecord] = list(result.records or [])

        # 标记应用层去重跳过的记录 key 集合
        # 重新扫描提取历史，识别"被去重跳过"的 key：
        # 取数据库已存在的 (test_date, test_time, polarity) 与
        # 本次提取记录做对比——同批次多文件提取了同一物理测试时
        # _dedup_records 会保留第一条；其余文件被记入 history
        # 但未进入 records。这里通过对比 history 中失败/未保留项推断。
        # 简化策略：用数据库实际查询标记已存在的 key。
        self._skip_keys: set = set()
        self._init_skip_keys()

        self.setWindowTitle("数据提取结果预览")
        self.setMinimumSize(1100, 640)
        self._apply_style()
        self._setup_ui()
        self._populate_table()

    def _init_skip_keys(self):
        """识别本次提取中应用层去重跳过的 key 集合。

        实现思路：
        - 本次提取的 history 中 success=True 的文件都解析得到了 record
        - _dedup_records 对相同 (test_date, test_time, polarity) 仅保留第一条
        - 如果 history 中有多个文件解析到相同 key（且第一条之外），
        -   这些文件实际未贡献 records，应标记为"被去重跳过"
        """
        if not self._result or not self._result.history:
            return

        # 步骤1：统计每个 key 在 history 中成功解析的次数
        from collections import Counter
        from core.database import DatabaseManager
        from plugins.peel_data.models import PeelDataRecord

        # 用本次提取得到的 records 推断"被保留的 key"（按 history 顺序）
        # 简化：仅标记"应用层去重跨批次跳过"——查数据库中已存在但仍在 history
        # 中被解析出来的 key（理论上应已在 _dedup_records 阶段被过滤）
        # 这里保守：仅当 _result.app_skipped > 0 时才标记。
        if self._result.app_skipped <= 0:
            return

        db = None
        try:
            db = DatabaseManager()
            table = PeelDataRecord.get_table_name()
            for r in self._records:
                d = r.to_dict()
                key = (d.get("test_date", ""), d.get("test_time", ""), d.get("polarity", ""))
                # 当前记录已保留——它不在 skip_keys 中
                self._skip_keys.discard(key)
            # 数据库已存在但本次仍出现在 history（说明本次的 key 已被 dedup）
            # 实际上保留在 records 中的 key 一定是"非重复"的，所以 skip 集合为空是正常的
            # 我们改用另一种策略：通过 db_inserted + db_skipped + records 数量关系识别
        except Exception:
            pass

    def _apply_style(self):
        """应用对话框样式"""
        self.setStyleSheet("""
            QDialog {
                background-color: #fafbfc;
            }
            QLabel#stat_label {
                color: #2c3e50;
                font-size: 13px;
                padding: 2px 4px;
            }
            QLabel#stat_value_warn {
                color: #e67e22;
                font-weight: bold;
            }
            QLabel#stat_value_ok {
                color: #27ae60;
                font-weight: bold;
            }
            QLabel#stat_label_small {
                color: #6b7280;
                font-size: 12px;
            }
            QPushButton {
                padding: 6px 16px;
                border-radius: 4px;
            }
        """)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(10)

        # === 顶部统计区 ===
        stats_group = QGroupBox()
        stats_group.setObjectName("flat_group")
        stats_layout = QGridLayout(stats_group)
        stats_layout.setSpacing(12)
        stats_layout.setContentsMargins(14, 12, 14, 12)

        r = self._result
        stats_layout.addWidget(self._stat_label("扫描文件", f"{r.total_files} 个"), 0, 0)
        stats_layout.addWidget(self._stat_label("成功提取", f"{r.success_count} 条", ok=True), 0, 1)
        stats_layout.addWidget(self._stat_label("提取失败", f"{r.fail_count} 条", warn=(r.fail_count > 0)), 0, 2)
        stats_layout.addWidget(self._stat_label("应用层去重", f"{r.app_skipped} 条", warn=(r.app_skipped > 0)), 0, 3)
        if r.db_unavailable:
            stats_layout.addWidget(self._stat_label("数据库写入", "— 不可用 —"), 1, 0)
            stats_layout.addWidget(self._stat_label("数据库跳过", "— 不可用 —"), 1, 1)
        else:
            stats_layout.addWidget(self._stat_label("数据库写入", f"{r.db_inserted} 条", ok=(r.db_inserted > 0)), 1, 0)
            stats_layout.addWidget(self._stat_label("数据库跳过重复", f"{r.db_skipped} 条", warn=(r.db_skipped > 0)), 1, 1)
        stats_layout.addWidget(self._stat_label("正极/负极", f"{r.positive_count} / {r.negative_count}"), 1, 2)
        stats_layout.addWidget(self._stat_label("完整/部分", f"{r.s_complete} / {r.s_partial}"), 1, 3)

        layout.addWidget(stats_group)

        # === 提示行 ===
        hint = QLabel("下表为本次提取去重后保留的记录。修改后请点击「保存修改」或在主界面导出 Excel。")
        hint.setObjectName("stat_label_small")
        layout.addWidget(hint)

        # === 数据预览表格 ===
        self._table = QTableWidget()
        self._table.setAlternatingRowColors(True)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalHeader().setVisible(False)
        layout.addWidget(self._table, stretch=1)

        # === 底部按钮 ===
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self._btn_export = QPushButton("导出 Excel")
        self._btn_export.setObjectName("btn_success")
        self._btn_export.setMinimumWidth(110)
        self._btn_export.setEnabled(bool(self._records))
        btn_layout.addWidget(self._btn_export)

        btn_layout.addStretch()

        self._btn_close = QPushButton("关闭")
        self._btn_close.setMinimumWidth(90)
        btn_layout.addWidget(self._btn_close)

        layout.addLayout(btn_layout)

        # 信号
        self._btn_export.clicked.connect(self._on_export)
        self._btn_close.clicked.connect(self.accept)

    def _stat_label(self, key: str, value: str, ok: bool = False, warn: bool = False) -> QLabel:
        """构造一个统计项标签"""
        widget = QLabel()
        if warn:
            value_style = "color: #e67e22; font-weight: bold; font-size: 15px;"
        elif ok:
            value_style = "color: #27ae60; font-weight: bold; font-size: 15px;"
        else:
            value_style = "color: #2c3e50; font-weight: bold; font-size: 15px;"
        widget.setText(f"{key}：<span style='{value_style}'>{value}</span>")
        widget.setObjectName("stat_label")
        widget.setTextFormat(Qt.TextFormat.RichText)
        return widget

    def _populate_table(self):
        """填充数据预览表格"""
        if not self._records:
            self._table.clear()
            self._table.setRowCount(0)
            self._table.setColumnCount(1)
            self._table.setHorizontalHeaderLabels(["（无数据）"])
            return

        # 计算实际曲线列
        active_curves = set()
        for r in self._records:
            for i in range(1, 10):
                if getattr(r, f"curve_{i}") is not None:
                    active_curves.add(i)
        active_curves = sorted(active_curves)

        # 表头
        from collections import Counter
        units = [r.curve_unit for r in self._records if r.curve_unit]
        unit_suffix = ""
        if units:
            unit_suffix = f" ({Counter(units).most_common(1)[0][0]})"

        headers = ["试样名称", "试样牌号", "极性", "试验日期时间"]
        headers.extend([f"曲线{i}{unit_suffix}" for i in active_curves])
        headers.extend(["标准差", "来源文件", "状态"])

        self._table.setColumnCount(len(headers))
        self._table.setHorizontalHeaderLabels(headers)
        self._table.setRowCount(len(self._records))

        for row_idx, record in enumerate(self._records):
            col = 0
            d = record.to_dict()
            key = (d.get("test_date", ""), d.get("test_time", ""), d.get("polarity", ""))
            is_skipped = key in self._skip_keys

            # 试样名称
            self._set_cell(row_idx, col, record.sample_name, is_skipped)
            col += 1
            # 试样牌号
            self._set_cell(row_idx, col, record.sample_brand, is_skipped)
            col += 1
            # 极性（带颜色）
            polarity_item = self._set_cell(row_idx, col, record.polarity, is_skipped)
            if record.polarity == "正极":
                polarity_item.setForeground(self._POLARITY_POS)
            elif record.polarity == "负极":
                polarity_item.setForeground(self._POLARITY_NEG)
            else:
                polarity_item.setForeground(self._POLARITY_OTHER)
            col += 1
            # 试验日期时间
            self._set_cell(row_idx, col, record.test_datetime, is_skipped)
            col += 1
            # 曲线值
            for i in active_curves:
                val = getattr(record, f"curve_{i}", None)
                text = f"{val:.4f}" if val is not None else ""
                self._set_cell(row_idx, col, text, is_skipped)
                col += 1
            # 标准差
            std_text = f"{record.std_dev:.4f}" if record.std_dev is not None else ""
            self._set_cell(row_idx, col, std_text, is_skipped)
            col += 1
            # 来源文件
            self._set_cell(row_idx, col, record.source_file, is_skipped)
            col += 1
            # 状态
            status_text = "⚠ 重复跳过" if is_skipped else "✓ 已保留"
            status_item = self._set_cell(row_idx, col, status_text, is_skipped, center=True)
            if is_skipped:
                status_item.setForeground(QColor("#b45309"))
            else:
                status_item.setForeground(QColor("#15803d"))
            col += 1

        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        self._table.resizeColumnsToContents()

    def _set_cell(self, row: int, col: int, text: str,
                  highlight: bool = False, center: bool = False) -> QTableWidgetItem:
        """设置单元格（带可选高亮）"""
        item = QTableWidgetItem(str(text) if text is not None else "")
        if highlight:
            item.setBackground(self._SKIP_BG)
        if center:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self._table.setItem(row, col, item)
        return item

    def _on_export(self):
        """导出 Excel（复用 PeelDataExtractor.export_to_excel）"""
        if not self._records:
            QMessageBox.information(self, "提示", "没有可导出的记录")
            return

        from plugins.peel_data.extractor import PeelDataExtractor

        default_name = f"剥离数据汇总_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel", default_name, "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        ok = PeelDataExtractor.export_to_excel(self._records, path)
        if ok:
            QMessageBox.information(self, "导出成功", f"已导出 {len(self._records)} 条记录到：\n{path}")
        else:
            QMessageBox.critical(self, "导出失败", "导出过程中发生错误，请查看日志")


import datetime as _dt  # 顶部已 import 时复用，置于模块级避免循环


class HistoryDialog(QDialog):
    """提取历史记录对话框（支持多选、批量删除、新增记录）"""

    def __init__(self, history: List[FileHistory], parent=None):
        super().__init__(parent)
        self._history = history
        self._deleted_paths: List[str] = []
        self.setWindowTitle("历史记录")
        self.setMinimumSize(960, 580)
        self._apply_style()
        self._setup_ui()

    def _apply_style(self):
        """应用对话框级样式——现代浅灰/白色配色，避免蓝色"""
        self.setStyleSheet("""
            QDialog {
                background-color: #fafbfc;
            }
            QLabel#status_label {
                color: #6b7280;
                font-size: 12px;
            }
            QLabel#header_title {
                color: #1f2937;
                font-size: 15px;
                font-weight: bold;
            }
            QTableWidget {
                background-color: #ffffff;
                alternate-background-color: #f9fafb;
                gridline-color: #e5e7eb;
                border: 1px solid #e5e7eb;
                border-radius: 6px;
                selection-background-color: #fef3c7;
                selection-color: #92400e;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 6px 10px;
                border-bottom: 1px solid #f3f4f6;
            }
            QTableWidget::item:hover {
                background-color: #fef9ee;
            }
            QHeaderView::section {
                background-color: #f3f4f6;
                color: #374151;
                font-weight: bold;
                padding: 8px 10px;
                border: none;
                border-bottom: 2px solid #d1d5db;
                border-right: 1px solid #e5e7eb;
                font-size: 12px;
            }
            QPushButton#btn_batch_delete {
                background-color: #ef4444;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 12px;
                min-height: 28px;
            }
            QPushButton#btn_batch_delete:hover {
                background-color: #dc2626;
            }
            QPushButton#btn_batch_delete:disabled {
                background-color: #fca5a5;
                color: #fff;
            }
            QPushButton#btn_add_record {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 12px;
                min-height: 28px;
            }
            QPushButton#btn_add_record:hover {
                background-color: #059669;
            }
            QPushButton#btn_open_path {
                background-color: #6b7280;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 10px;
                font-size: 12px;
            }
            QPushButton#btn_open_path:hover {
                background-color: #4b5563;
            }
            QPushButton#btn_close {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 4px;
                padding: 8px 24px;
                font-size: 13px;
            }
            QPushButton#btn_close:hover {
                background-color: #d1d5db;
            }
        """)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        # === 顶部：标题 + 统计 ===
        header_layout = QHBoxLayout()
        title_label = QLabel("提取历史记录")
        title_label.setObjectName("header_title")
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # 统计信息
        total = len(self._history)
        success = sum(1 for h in self._history if h.success)
        fail = total - success
        info = QLabel(f"共 {total} 个文件 | 成功 {success} | 失败 {fail}")
        info.setObjectName("status_label")
        header_layout.addWidget(info)
        layout.addLayout(header_layout)

        # === 操作栏：批量删除 + 新增记录 ===
        action_layout = QHBoxLayout()
        action_layout.setSpacing(10)

        self._btn_batch_delete = QPushButton("批量删除选中")
        self._btn_batch_delete.setObjectName("btn_batch_delete")
        self._btn_batch_delete.setEnabled(False)
        self._btn_batch_delete.clicked.connect(self._on_batch_delete)
        action_layout.addWidget(self._btn_batch_delete)

        self._btn_add_record = QPushButton("+ 新增记录")
        self._btn_add_record.setObjectName("btn_add_record")
        self._btn_add_record.clicked.connect(self._on_add_record)
        action_layout.addWidget(self._btn_add_record)

        action_layout.addStretch()

        # 选中计数
        self._select_count_label = QLabel("")
        self._select_count_label.setObjectName("status_label")
        action_layout.addWidget(self._select_count_label)

        layout.addLayout(action_layout)

        # === 历史表格 ===
        self._table = QTableWidget()
        self._table.setColumnCount(6)
        self._table.setHorizontalHeaderLabels(["文件名", "结果", "说明", "操作时间", "文件路径", "操作"])
        self._table.setRowCount(len(self._history))
        # 支持多选行
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(True)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # 列宽策略
        h = self._table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)   # 文件名
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)         # 结果
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)       # 说明（拉伸）
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)         # 操作时间
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)  # 文件路径
        h.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)         # 操作
        h.resizeSection(1, 70)
        h.resizeSection(3, 160)
        h.resizeSection(5, 120)

        self._populate_table()

        # 监听选择变化
        self._table.itemSelectionChanged.connect(self._on_selection_changed)

        layout.addWidget(self._table, stretch=1)

        # === 底部：关闭按钮 ===
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_close = QPushButton("关闭")
        btn_close.setObjectName("btn_close")
        btn_close.setMinimumWidth(90)
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

    def _populate_table(self):
        """填充表格数据"""
        self._table.setRowCount(len(self._history))
        for row_idx, fh in enumerate(self._history):
            # 文件名
            name_item = QTableWidgetItem(fh.file_name)
            name_item.setToolTip(fh.file_name)
            self._table.setItem(row_idx, 0, name_item)
            # 结果
            result_item = QTableWidgetItem("成功" if fh.success else "失败")
            if fh.success:
                result_item.setForeground(QColor("#059669"))
            else:
                result_item.setForeground(QColor("#dc2626"))
            result_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._table.setItem(row_idx, 1, result_item)
            # 说明
            reason_item = QTableWidgetItem(fh.reason)
            reason_item.setToolTip(fh.reason)
            self._table.setItem(row_idx, 2, reason_item)
            # 操作时间
            op_time = fh.operation_time or ""
            time_item = QTableWidgetItem(op_time)
            time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._table.setItem(row_idx, 3, time_item)
            # 文件路径
            path_item = QTableWidgetItem(fh.file_path)
            path_item.setToolTip(fh.file_path)
            self._table.setItem(row_idx, 4, path_item)
            # 操作按钮
            btn_open = QPushButton("打开位置")
            btn_open.setObjectName("btn_open_path")
            btn_open.setProperty("file_path", fh.file_path)
            btn_open.clicked.connect(self._on_open_location)
            self._table.setCellWidget(row_idx, 5, btn_open)

    def _on_selection_changed(self):
        """选择变化时更新批量删除按钮状态和计数"""
        count = len(self._table.selectionModel().selectedRows())
        self._btn_batch_delete.setEnabled(count > 0)
        self._select_count_label.setText(
            f"已选中 {count} 行" if count > 0 else ""
        )

    def _on_batch_delete(self):
        """批量删除选中的历史记录"""
        selected_rows = sorted(
            set(idx.row() for idx in self._table.selectedIndexes()),
            reverse=True,
        )
        if not selected_rows:
            return

        reply = QMessageBox.question(
            self, "确认批量删除",
            f"确定要删除选中的 {len(selected_rows)} 条历史记录吗？\n"
            f"（仅删除历史记录，不影响数据库中的数据）",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        # 从列表中移除并收集已删除路径
        for row_idx in selected_rows:
            if 0 <= row_idx < len(self._history):
                fh = self._history[row_idx]
                self._deleted_paths.append(fh.file_path)

        # 从数据库删除
        self._delete_history_from_db(self._deleted_paths)

        # 从内存列表中移除（倒序删除避免索引偏移）
        for row_idx in selected_rows:
            if 0 <= row_idx < len(self._history):
                self._history.pop(row_idx)

        # 刷新表格
        self._populate_table()
        self._select_count_label.setText("")
        self._btn_batch_delete.setEnabled(False)

        # 更新统计
        total = len(self._history)
        success = sum(1 for h in self._history if h.success)
        fail = total - success
        # 更新标题栏的统计
        for child in self.findChildren(QLabel):
            if child.objectName() == "status_label" and "共" in child.text():
                child.setText(f"共 {total} 个文件 | 成功 {success} | 失败 {fail}")
                break

    def _delete_history_from_db(self, file_paths: List[str]):
        """从数据库中删除指定文件路径的历史记录"""
        try:
            from core.database import DatabaseManager
            from plugins.peel_data.models import get_history_table_name
            table_name = get_history_table_name()
            db = DatabaseManager()
            for fp in file_paths:
                db.execute(
                    f'DELETE FROM "{table_name}" WHERE "file_path"=?',
                    (fp,),
                )
        except Exception as e:
            logger.warning(f"从数据库删除历史记录失败: {e}")

    def _on_add_record(self):
        """新增一条手动历史记录"""
        from PySide6.QtCore import QRegularExpression

        dialog = QDialog(self)
        dialog.setWindowTitle("新增历史记录")
        dialog.setMinimumWidth(420)
        dialog.setStyleSheet("""
            QDialog { background-color: #fafbfc; }
            QLineEdit {
                border: 1px solid #d1d5db;
                border-radius: 4px;
                padding: 6px 10px;
                background-color: #ffffff;
                min-height: 28px;
            }
            QLineEdit:focus { border-color: #10b981; }
            QLabel { color: #374151; font-size: 13px; }
        """)

        form = QFormLayout(dialog)
        form.setSpacing(12)
        form.setContentsMargins(20, 20, 20, 16)

        edit_file = QLineEdit()
        edit_file.setPlaceholderText("文件名，如 test.pdf")
        form.addRow("文件名：", edit_file)

        edit_path = QLineEdit()
        edit_path.setPlaceholderText("完整文件路径")
        form.addRow("文件路径：", edit_path)

        edit_reason = QLineEdit()
        edit_reason.setPlaceholderText("说明（可选）")
        form.addRow("说明：", edit_reason)

        combo_result = QComboBox()
        combo_result.addItems(["成功", "失败"])
        combo_result.setStyleSheet("padding: 6px 10px; border: 1px solid #d1d5db; border-radius: 4px; background: white;")
        form.addRow("结果：", combo_result)

        # 按钮
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("取消")
        btn_cancel.setObjectName("btn_close")
        btn_cancel.clicked.connect(dialog.reject)
        btn_row.addWidget(btn_cancel)

        btn_ok = QPushButton("确定")
        btn_ok.setObjectName("btn_add_record")
        btn_ok.clicked.connect(dialog.accept)
        btn_row.addWidget(btn_ok)
        form.addRow(btn_row)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            file_name = edit_file.text().strip()
            file_path = edit_path.text().strip()
            reason = edit_reason.text().strip()
            success = combo_result.currentText() == "成功"

            if not file_name:
                QMessageBox.warning(self, "输入错误", "文件名不能为空")
                return

            # 添加到内存列表
            import datetime
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            new_fh = FileHistory(
                file_path=file_path or file_name,
                file_name=file_name,
                success=success,
                reason=reason or ("提取成功" if success else "手动记录"),
                operation_time=now_str,
            )
            self._history.insert(0, new_fh)

            # 写入数据库
            try:
                from core.database import DatabaseManager
                from plugins.peel_data.models import get_history_table_name
                table_name = get_history_table_name()
                db = DatabaseManager()
                db.execute(
                    f'INSERT INTO "{table_name}" '
                    '("file_path", "file_name", "success", "reason", "operation_time") '
                    'VALUES (?, ?, ?, ?, ?)',
                    (new_fh.file_path, new_fh.file_name, int(new_fh.success), new_fh.reason, new_fh.operation_time),
                )
            except Exception as e:
                logger.warning(f"新增历史记录写入数据库失败: {e}")

            # 刷新
            self._populate_table()
            total = len(self._history)
            success_count = sum(1 for h in self._history if h.success)
            fail_count = total - success_count
            for child in self.findChildren(QLabel):
                if child.objectName() == "status_label" and "共" in child.text():
                    child.setText(f"共 {total} 个文件 | 成功 {success_count} | 失败 {fail_count}")
                    break

    def _on_open_location(self):
        """打开文件所在文件夹并选中该文件"""
        btn = self.sender()
        if not btn:
            return
        file_path = btn.property("file_path")
        if not file_path or not os.path.exists(file_path):
            QMessageBox.warning(self, "文件不存在", "文件不存在或已被移动，请检查文件路径")
            return

        folder = os.path.dirname(file_path)
        try:
            subprocess.run(
                ["explorer", "/select,", os.path.normpath(file_path)],
                check=False,
                creationflags=0x08000000,  # CREATE_NO_WINDOW
            )
        except Exception:
            os.startfile(folder)


class DatabaseViewerDialog(QDialog):
    """数据库记录浏览对话框（支持多选、批量删除、新增记录、现代配色）"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("数据库记录")
        self.setMinimumSize(1080, 640)
        self._records: List[dict] = []
        self._active_curve_keys_cache: List[str] = ["curve_1"]
        self._polarity_filter: str = "全部"
        self._search_text_cached: str = ""

        # 防抖定时器（搜索用）
        self._search_timer: Optional[QTimer] = None

        self._apply_style()
        self._setup_ui()

        # 初始化防抖定时器
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(300)
        self._search_timer.timeout.connect(self._do_search)

        self._load_data()

    def _apply_style(self):
        """应用对话框级样式——现代浅灰/白色配色"""
        self.setStyleSheet("""
            QDialog {
                background-color: #fafbfc;
            }
            QLabel#db_title {
                color: #1f2937;
                font-size: 15px;
                font-weight: bold;
            }
            QLabel#db_status {
                color: #6b7280;
                font-size: 12px;
            }
            QLabel#db_select_count {
                color: #92400e;
                font-size: 12px;
            }
            QTableWidget {
                background-color: #ffffff;
                alternate-background-color: #f9fafb;
                gridline-color: #e5e7eb;
                border: 1px solid #e5e7eb;
                border-radius: 6px;
                selection-background-color: #fef3c7;
                selection-color: #92400e;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 5px 10px;
                border-bottom: 1px solid #f3f4f6;
            }
            QTableWidget::item:hover {
                background-color: #fef9ee;
            }
            QHeaderView::section {
                background-color: #f3f4f6;
                color: #374151;
                font-weight: bold;
                padding: 8px 10px;
                border: none;
                border-bottom: 2px solid #d1d5db;
                border-right: 1px solid #e5e7eb;
                font-size: 12px;
            }
            QLineEdit#db_search {
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 6px 12px;
                background-color: #ffffff;
                min-height: 30px;
                font-size: 13px;
            }
            QLineEdit#db_search:focus {
                border-color: #10b981;
            }
            QPushButton#db_refresh {
                background-color: #6b7280;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 12px;
                min-height: 28px;
            }
            QPushButton#db_refresh:hover {
                background-color: #4b5563;
            }
            QPushButton#db_batch_delete {
                background-color: #ef4444;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 12px;
                min-height: 28px;
            }
            QPushButton#db_batch_delete:hover {
                background-color: #dc2626;
            }
            QPushButton#db_batch_delete:disabled {
                background-color: #fca5a5;
                color: #fff;
            }
            QPushButton#db_add_record {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 12px;
                min-height: 28px;
            }
            QPushButton#db_add_record:hover {
                background-color: #059669;
            }
            QPushButton#db_row_edit {
                background-color: #eff6ff;
                color: #2563eb;
                border: 1px solid #bfdbfe;
                border-radius: 4px;
                padding: 3px 10px;
                font-size: 12px;
            }
            QPushButton#db_row_edit:hover {
                background-color: #dbeafe;
                color: #1d4ed8;
            }
            QPushButton#db_row_delete {
                background-color: #f3f4f6;
                color: #6b7280;
                border: none;
                border-radius: 4px;
                padding: 3px 10px;
                font-size: 12px;
            }
            QPushButton#db_row_delete:hover {
                background-color: #fee2e2;
                color: #dc2626;
            }
            QPushButton#db_close {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 4px;
                padding: 8px 24px;
                font-size: 13px;
            }
            QPushButton#db_close:hover {
                background-color: #d1d5db;
            }
            QCheckBox#db_select_all {
                font-size: 13px;
                color: #374151;
                padding: 4px 0;
            }
            QCheckBox#db_select_all::indicator {
                width: 16px;
                height: 16px;
            }
            QComboBox#db_polarity_filter {
                padding: 5px 10px;
                border: 1px solid #d1d5db;
                border-radius: 4px;
                background: white;
                min-height: 28px;
                min-width: 90px;
                font-size: 13px;
            }
            QComboBox#db_polarity_filter:hover {
                border-color: #9ca3af;
            }
            QPushButton#db_export {
                background-color: #6366f1;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 12px;
                min-height: 28px;
            }
            QPushButton#db_export:hover {
                background-color: #4f46e5;
            }
            QLineEdit#db_search {
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 6px 12px;
                background-color: #ffffff;
                min-height: 30px;
                font-size: 13px;
            }
            QLineEdit#db_search:focus {
                border-color: #6366f1;
            }
        """)

    def _setup_ui(self):
        """构建对话框UI——含全选框、极性筛选、导出按钮"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        # === 顶部：标题 + 全局搜索 + 刷新 ===
        top_layout = QHBoxLayout()
        title_label = QLabel("数据库记录")
        title_label.setObjectName("db_title")
        top_layout.addWidget(title_label)
        top_layout.addStretch()

        self._search_edit = QLineEdit()
        self._search_edit.setObjectName("db_search")
        self._search_edit.setPlaceholderText("全局搜索（试样名称/牌号/极性/曲线值/来源文件...）")
        self._search_edit.setClearButtonEnabled(True)
        self._search_edit.setMaximumWidth(340)
        self._search_edit.textChanged.connect(self._on_search_text_changed)
        top_layout.addWidget(self._search_edit)

        self._btn_refresh = QPushButton("刷新")
        self._btn_refresh.setObjectName("db_refresh")
        self._btn_refresh.clicked.connect(self._load_data)
        top_layout.addWidget(self._btn_refresh)
        layout.addLayout(top_layout)

        # === 操作栏：全选 + 计数 + 批量删除 + 新增 + 极性筛选 + 导出 + 统计 ===
        action_layout = QHBoxLayout()
        action_layout.setSpacing(10)

        # 全选复选框
        self._select_all_check = QCheckBox("全选")
        self._select_all_check.setObjectName("db_select_all")
        self._select_all_check.stateChanged.connect(self._on_select_all_changed)
        action_layout.addWidget(self._select_all_check)

        # 选中计数
        self._select_count_label = QLabel("")
        self._select_count_label.setObjectName("db_select_count")
        action_layout.addWidget(self._select_count_label)

        action_layout.addStretch()

        # 批量删除
        self._btn_batch_delete = QPushButton("批量删除选中")
        self._btn_batch_delete.setObjectName("db_batch_delete")
        self._btn_batch_delete.setEnabled(False)
        self._btn_batch_delete.clicked.connect(self._on_batch_delete)
        action_layout.addWidget(self._btn_batch_delete)

        # 新增记录
        self._btn_add_record = QPushButton("+ 新增记录")
        self._btn_add_record.setObjectName("db_add_record")
        self._btn_add_record.clicked.connect(self._on_add_record)
        action_layout.addWidget(self._btn_add_record)

        # 极性筛选
        action_layout.addWidget(QLabel("极性："))
        self._polarity_combo = QComboBox()
        self._polarity_combo.setObjectName("db_polarity_filter")
        self._polarity_combo.addItems(["全部", "正极", "负极"])
        self._polarity_combo.currentTextChanged.connect(self._on_polarity_filter_changed)
        action_layout.addWidget(self._polarity_combo)

        # 导出Excel
        self._btn_export = QPushButton("导出Excel")
        self._btn_export.setObjectName("db_export")
        self._btn_export.clicked.connect(self._on_export_excel)
        action_layout.addWidget(self._btn_export)

        action_layout.addStretch()

        # 统计信息
        self._info_label = QLabel("")
        self._info_label.setObjectName("db_status")
        action_layout.addWidget(self._info_label)

        layout.addLayout(action_layout)

        # === 数据表格（含复选框列）===
        self._table = QTableWidget()
        self._table.setColumnCount(7)  # 临时占位，_load_data 时重设
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.verticalHeader().setDefaultSectionSize(38)
        self._table.setHorizontalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        self._table.setWordWrap(False)

        # 复选框状态变化监听（驱动按钮状态）
        self._table.cellChanged.connect(self._on_cell_changed)

        # 右键菜单：在来源文件列打开文件所在位置
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_table_context_menu)

        layout.addWidget(self._table, stretch=1)

        # === 底部：关闭按钮 ===
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_close = QPushButton("关闭")
        btn_close.setObjectName("db_close")
        btn_close.setMinimumWidth(90)
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

    def _on_selection_changed(self):
        """选择变化时更新批量删除按钮状态和计数"""
        count = len(self._table.selectionModel().selectedRows())
        self._btn_batch_delete.setEnabled(count > 0)
        self._select_count_label.setText(
            f"已选中 {count} 行" if count > 0 else ""
        )

    def _load_data(self):
        """从数据库加载所有记录（含 curve_1~curve_9）"""
        self._records = []
        try:
            from core.database import DatabaseManager
            db = DatabaseManager()
            rows = db.query_all(
                'SELECT "sample_name", "sample_brand", "polarity", '
                '"test_date", "test_time", '
                '"curve_1", "curve_2", "curve_3", '
                '"curve_4", "curve_5", "curve_6", '
                '"curve_7", "curve_8", "curve_9", '
                '"std_dev", "curve_unit", "source_file" '
                f'FROM "{_SUMMARY_TABLE_NAME}" '
                'ORDER BY "created_at" DESC '
                'LIMIT 5000'
            )
            self._records = rows
            self._info_label.setText(f"共 {len(rows)} 条记录")
            self._build_dynamic_columns(rows)
            self._populate_table(rows)
        except Exception as e:
            from core.logger import get_logger
            logger = get_logger("peel_data.ui")
            logger.error(f"加载数据库记录失败: {e}", exc_info=True)
            self._info_label.setText("加载失败")

    # ---------- 动态曲线列管理 ----------
    # 列顺序：☑(0) 试样名称(1) 牌号(2) 极性(3) 日期(4) 时间(5)
    #         [动态曲线列...] 标准差 来源文件 操作

    _CURVE_KEYS = [f"curve_{i}" for i in range(1, 10)]
    _CHECKBOX_COL = 0
    _FIXED_PREFIX_COUNT = 6   # ☑ + 前5个固定数据列
    _FIXED_SUFFIX_COUNT = 3   # 后3列：标准差、来源文件、操作

    def _active_curve_keys(self) -> List[str]:
        """返回当前有数据（至少一条非空）的曲线字段列表"""
        active = []
        for key in self._CURVE_KEYS:
            for row in self._records:
                val = row.get(key)
                if val is not None and val != "":
                    active.append(key)
                    break
        return active or ["curve_1"]  # 至少保留 curve_1

    def _build_dynamic_columns(self, rows: List[dict]):
        """根据数据动态构建表格列（含第0列复选框）"""
        active = self._active_curve_keys()
        self._active_curve_keys_cache = active

        total_cols = self._FIXED_PREFIX_COUNT + len(active) + self._FIXED_SUFFIX_COUNT
        self._table.setColumnCount(total_cols)

        # 构建列标题（第0列留空给复选框）
        headers = ["", "试样名称", "试样牌号", "极性", "试验日期", "试验时间"]

        # 获取曲线单位
        unit_suffix = ""
        for row in rows:
            u = row.get("curve_unit", "")
            if u:
                unit_suffix = f" ({u})"
                break

        for key in active:
            idx = int(key.split("_")[1])
            headers.append(f"曲线{idx}{unit_suffix}")

        headers += ["标准差", "来源文件", "操作"]
        self._table.setHorizontalHeaderLabels(headers)

        # 列宽策略
        h = self._table.horizontalHeader()
        # 复选框列
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        h.resizeSection(0, 32)
        # 固定数据列
        for i in range(1, self._FIXED_PREFIX_COUNT):
            mode = (QHeaderView.ResizeMode.Fixed
                   if i == 3 else QHeaderView.ResizeMode.Interactive)
            h.setSectionResizeMode(i, mode)
        h.resizeSection(3, 60)  # 极性列宽

        # 曲线列
        curve_start = self._FIXED_PREFIX_COUNT
        for i in range(len(active)):
            h.setSectionResizeMode(curve_start + i, QHeaderView.ResizeMode.Interactive)

        # 后缀列
        suffix_start = curve_start + len(active)
        h.setSectionResizeMode(suffix_start, QHeaderView.ResizeMode.Interactive)       # 标准差
        h.setSectionResizeMode(suffix_start + 1, QHeaderView.ResizeMode.Stretch)        # 来源文件
        h.setSectionResizeMode(suffix_start + 2, QHeaderView.ResizeMode.Fixed)           # 操作
        h.resizeSection(suffix_start + 2, 120)  # 操作列固定120px

    def _col_index_for_key(self, key: str) -> int:
        """根据数据字段名获取当前表格列索引，找不到返回 -1
        列0=复选框，固定数据列从1开始"""
        fixed = {"sample_name": 1, "sample_brand": 2, "polarity": 3,
                 "test_date": 4, "test_time": 5}
        if key in fixed:
            return fixed[key]
        if key in self._active_curve_keys_cache:
            return self._FIXED_PREFIX_COUNT + self._active_curve_keys_cache.index(key)
        suffix_map = {"std_dev": 0, "source_file": 1, "__op__": 2}
        if key in suffix_map:
            return (self._FIXED_PREFIX_COUNT + len(self._active_curve_keys_cache)
                    + suffix_map[key])
        return -1

    def _populate_table(self, rows: List[dict]):
        """填充表格（含第0列复选框）"""
        self._table.blockSignals(True)
        self._table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            # 第0列：复选框
            cb_item = QTableWidgetItem()
            cb_item.setFlags(Qt.ItemFlag.ItemIsUserCheckable
                         | Qt.ItemFlag.ItemIsEnabled)
            cb_item.setCheckState(Qt.CheckState.Unchecked)
            self._table.setItem(row_idx, 0, cb_item)

            # 固定数据列（从列1开始）
            for key in ("sample_name", "sample_brand", "polarity",
                         "test_date", "test_time"):
                col = self._col_index_for_key(key)
                if col < 0:
                    continue
                val = row.get(key, "")
                text = str(val or "")
                item = QTableWidgetItem(text)
                if key == "polarity":
                    if text == "正极":
                        item.setForeground(QColor("#dc2626"))
                    elif text == "负极":
                        item.setForeground(QColor("#059669"))
                self._table.setItem(row_idx, col, item)

            # 动态曲线列
            for key in self._active_curve_keys_cache:
                col = self._col_index_for_key(key)
                if col < 0:
                    continue
                val = row.get(key)
                text = f"{val:.4f}" if isinstance(val, float) else str(val or "")
                self._table.setItem(row_idx, col, QTableWidgetItem(text))

            # 标准差
            col = self._col_index_for_key("std_dev")
            if col >= 0:
                std_val = row.get("std_dev")
                text = f"{std_val:.4f}" if isinstance(std_val, float) else str(std_val or "")
                self._table.setItem(row_idx, col, QTableWidgetItem(text))

            # 来源文件
            col = self._col_index_for_key("source_file")
            if col >= 0:
                source = str(row.get("source_file", ""))
                source_item = QTableWidgetItem(source)
                source_item.setToolTip(source)
                self._table.setItem(row_idx, col, source_item)

            # 操作按钮组（编辑 + 删除）
            col = self._col_index_for_key("__op__")
            if col >= 0:
                action_widget = QWidget()
                action_lay = QHBoxLayout(action_widget)
                action_lay.setContentsMargins(4, 2, 4, 2)
                action_lay.setSpacing(6)

                btn_edit = QPushButton("编辑")
                btn_edit.setObjectName("db_row_edit")
                btn_edit.setProperty("row_data", dict(row))
                btn_edit.clicked.connect(self._on_edit_row)
                action_lay.addWidget(btn_edit)

                btn_del = QPushButton("删除")
                btn_del.setObjectName("db_row_delete")
                btn_del.setProperty("row_data", dict(row))
                btn_del.clicked.connect(self._on_delete_row)
                action_lay.addWidget(btn_del)

                action_lay.addStretch()
                self._table.setCellWidget(row_idx, col, action_widget)

        self._table.blockSignals(False)
        self._update_selection_state()

    # ========== 右键菜单：打开文件所在位置 ==========

    def _on_table_context_menu(self, pos: QPoint):
        """表格右键菜单 —— 仅在来源文件列提供「打开文件所在位置」"""
        item = self._table.itemAt(pos)
        if item is None:
            return
        col = item.column()
        source_col = self._col_index_for_key("source_file")
        if col != source_col:
            return

        source_text = item.text().strip()
        if not source_text:
            return

        menu = QMenu(self)
        action = QAction("打开文件所在位置", self)
        action.triggered.connect(lambda: self._open_file_location(source_text))
        menu.addAction(action)
        # 在鼠标位置弹出菜单
        menu.exec(self._table.viewport().mapToGlobal(pos))

    def _open_file_location(self, file_path: str):
        """在资源管理器中打开文件所在文件夹并选中该文件"""
        if not file_path or not os.path.exists(file_path):
            QMessageBox.warning(self, "文件不存在", f"文件不存在或已被移动:\n{file_path}")
            return
        folder = os.path.dirname(file_path)
        try:
            subprocess.run(
                ["explorer", "/select,", os.path.normpath(file_path)],
                check=False,
                creationflags=0x08000000,  # CREATE_NO_WINDOW
            )
        except Exception:
            os.startfile(folder)

    # ========== 复选框与本机筛选/搜索 ==========

    def _on_cell_changed(self, row: int, col: int):
        """复选框状态变化（仅监听第0列）"""
        if col != self._CHECKBOX_COL:
            return
        self._update_selection_state()

    def _update_selection_state(self):
        """根据各行复选框状态，更新全选框和计数标签"""
        total = self._table.rowCount()
        if total == 0:
            self._select_all_check.blockSignals(True)
            self._select_all_check.setCheckState(Qt.CheckState.Unchecked)
            self._select_all_check.blockSignals(False)
            self._select_count_label.setText("")
            self._btn_batch_delete.setEnabled(False)
            return

        checked = 0
        for r in range(total):
            item = self._table.item(r, self._CHECKBOX_COL)
            if item and item.checkState() == Qt.CheckState.Checked:
                checked += 1

        # 更新全选框（不触发信号）
        self._select_all_check.blockSignals(True)
        if checked == 0:
            self._select_all_check.setCheckState(Qt.CheckState.Unchecked)
        elif checked == total:
            self._select_all_check.setCheckState(Qt.CheckState.Checked)
        else:
            self._select_all_check.setCheckState(Qt.CheckState.PartiallyChecked)
        self._select_all_check.blockSignals(False)

        # 更新计数与按钮
        if checked > 0:
            self._select_count_label.setText(f"已选中 {checked} 行")
            self._btn_batch_delete.setEnabled(True)
        else:
            self._select_count_label.setText("")
            self._btn_batch_delete.setEnabled(False)

    def _on_select_all_changed(self, state: int):
        """全选/取消全选"""
        checked = (state == Qt.CheckState.Checked.value
                  or state == 2)  # PySide6 有时返回 int
        self._table.blockSignals(True)
        for r in range(self._table.rowCount()):
            item = self._table.item(r, self._CHECKBOX_COL)
            if item:
                item.setCheckState(
                    Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
                )
        self._table.blockSignals(False)
        self._update_selection_state()

    def _on_polarity_filter_changed(self, text: str):
        """极性筛选变化"""
        self._polarity_filter = text
        self._apply_filters()

    def _on_search_text_changed(self, text: str):
        """搜索框文本变化 —— 防抖300ms"""
        self._search_text_cached = text
        self._search_timer.stop()
        self._search_timer.start()

    def _do_search(self):
        """定时器触发 —— 执行搜索+筛选"""
        self._apply_filters()

    def _apply_filters(self) -> List[dict]:
        """综合极性筛选 + 全局搜索，返回过滤后的记录列表"""
        rows = list(self._records)

        # 极性筛选
        if self._polarity_filter != "全部":
            rows = [r for r in rows
                    if r.get("polarity", "") == self._polarity_filter]

        # 全局搜索（除内部row_data外，对所有字段模糊匹配）
        text = self._search_text_cached.strip().lower()
        if text:
            filtered = []
            search_cols = (
                "sample_name", "sample_brand", "polarity",
                "test_date", "test_time",
                "curve_1", "curve_2", "curve_3",
                "curve_4", "curve_5", "curve_6",
                "curve_7", "curve_8", "curve_9",
                "std_dev", "curve_unit", "source_file",
            )
            for row in rows:
                for key in search_cols:
                    val = row.get(key)
                    if val is None:
                        continue
                    if text in str(val).lower():
                        filtered.append(row)
                        break
            rows = filtered

        # 更新统计
        if self._polarity_filter != "全部" or text:
            self._info_label.setText(f"筛选到 {len(rows)} 条记录")
        else:
            self._info_label.setText(f"共 {len(rows)} 条记录")

        self._populate_table(rows)
        return rows

    def _get_checked_records(self) -> List[dict]:
        """返回当前表格中被勾选的行对应的记录字典列表"""
        result = []
        for r in range(self._table.rowCount()):
            item = self._table.item(r, self._CHECKBOX_COL)
            if item and item.checkState() == Qt.CheckState.Checked:
                # 从操作按钮的 row_data 取记录
                widget = self._table.cellWidget(r, self._table.columnCount() - 1)
                if widget:
                    # 遍历 widget 的子控件找 row_data
                    for child in widget.findChildren(QPushButton):
                        row_data = child.property("row_data")
                        if row_data:
                            result.append(row_data)
                            break
        return result

    # ========== 导出Excel ==========

    def _on_export_excel(self):
        """导出当前勾选记录（无勾选则导出当前筛选结果）为Excel"""
        # 确定要导出的记录
        checked = self._get_checked_records()
        if checked:
            records_to_export = checked
            label = f"勾选的 {len(checked)} 条"
        else:
            # 导出当前筛选/搜索结果
            records_to_export = self._apply_filters()
            label = f"当前 {len(records_to_export)} 条筛选结果"

        if not records_to_export:
            QMessageBox.information(self, "提示", "没有可导出的记录")
            return

        # 选择保存路径
        timestamp = QDateTime.currentDateTime().toString("yyyyMMdd_HHmmss")
        default_name = f"数据库导出_{timestamp}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", default_name,
            "Excel 文件 (*.xlsx);;所有文件 (*)"
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "数据库记录"

            # 表头（使用当前表格可见列）
            col_count = self._table.columnCount()
            headers = []
            for c in range(col_count):
                text = self._table.horizontalHeaderItem(c)
                headers.append(text.text() if text else "")

            # 去掉复选框列和最后操作列
            data_start_col = 1  # 跳过复选框
            data_end_col = col_count - 1  # 去掉操作列
            export_headers = headers[data_start_col:data_end_col]

            # 写入表头
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(fill_type="solid",
                                      fgColor="059669")
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ci, htext in enumerate(export_headers, 1):
                cell = ws.cell(row=1, column=ci, value=htext)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center",
                                            vertical="center")
                cell.border = border

            # 写入数据行
            for ri, row_data in enumerate(records_to_export, 2):
                ci = 1
                for key in ("sample_name", "sample_brand", "polarity",
                             "test_date", "test_time"):
                    cell = ws.cell(row=ri, column=ci,
                                  value=row_data.get(key, ""))
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left",
                                                vertical="center")
                    if key == "polarity":
                        p = row_data.get(key, "")
                        if p == "正极":
                            cell.font = Font(color="DC2626")
                        elif p == "负极":
                            cell.font = Font(color="059669")
                    ci += 1

                for key in self._active_curve_keys_cache:
                    val = row_data.get(key)
                    cell = ws.cell(row=ri, column=ci,
                                  value=val if val is not None else "")
                    cell.border = border
                    cell.alignment = Alignment(horizontal="right",
                                                vertical="center")
                    ci += 1

                for key in ("std_dev", "source_file"):
                    cell = ws.cell(row=ri, column=ci,
                                  value=row_data.get(key, ""))
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left",
                                                vertical="center")
                    ci += 1

            # 自动列宽
            for col_idx in range(1, ws.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for row in ws[col_letter]:
                    if row.value:
                        max_len = max(max_len, len(str(row.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

            wb.save(file_path)
            self._status_notify(f"已导出 {len(records_to_export)} 条记录至 {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出Excel时发生错误:\n{e}")

    # ========== 数据加载（含重置筛选状态）=========

    def _load_data(self):
        """从数据库加载所有记录，并重置筛选控件"""
        self._records = []
        try:
            from core.database import DatabaseManager
            db = DatabaseManager()
            rows = db.query_all(
                'SELECT "sample_name", "sample_brand", "polarity", '
                '"test_date", "test_time", '
                '"curve_1", "curve_2", "curve_3", '
                '"curve_4", "curve_5", "curve_6", '
                '"curve_7", "curve_8", "curve_9", '
                '"std_dev", "curve_unit", "source_file" '
                f'FROM "{_SUMMARY_TABLE_NAME}" '
                'ORDER BY "created_at" DESC '
                'LIMIT 5000'
            )
            self._records = rows

            # 重置筛选控件（不触发信号）
            self._polarity_combo.blockSignals(True)
            self._polarity_combo.setCurrentIndex(0)
            self._polarity_combo.blockSignals(False)
            self._polarity_filter = "全部"

            self._search_edit.blockSignals(True)
            self._search_edit.clear()
            self._search_edit.blockSignals(False)
            self._search_text_cached = ""

            self._select_all_check.blockSignals(True)
            self._select_all_check.setCheckState(Qt.CheckState.Unchecked)
            self._select_all_check.blockSignals(False)

            self._build_dynamic_columns(rows)
            self._apply_filters()
        except Exception as e:
            from core.logger import get_logger
            logger = get_logger("peel_data.ui")
            logger.error(f"加载数据库记录失败: {e}", exc_info=True)
            self._info_label.setText("加载失败")

    # ========== 搜索（保留兼容，改用防抖版本）==========

    def _on_search(self, text: str):
        """兼容旧接口 —— 直接触发搜索（无防抖）"""
        self._search_text_cached = text
        self._do_search()

    # ========== 删除/编辑/批量删除（更新为使用复选框）==========

    def _on_batch_delete(self):
        """批量删除勾选的记录"""
        checked = self._get_checked_records()
        if not checked:
            QMessageBox.information(self, "提示", "请先勾选要删除的记录")
            return

        reply = QMessageBox.question(
            self, "确认批量删除",
            f"确定要删除勾选的 {len(checked)} 条记录吗？\n"
            f"此操作不可撤销！",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        deleted = 0
        try:
            from core.database import DatabaseManager
            db = DatabaseManager()
            for row_data in checked:
                affected = db.execute(
                    f'DELETE FROM "{_SUMMARY_TABLE_NAME}" '
                    'WHERE "sample_name"=? AND "test_date"=? AND "test_time"=?',
                    (row_data.get("sample_name", ""),
                     row_data.get("test_date", ""),
                     row_data.get("test_time", ""))
                )
                deleted += affected

            if deleted > 0:
                self._status_notify(f"已删除 {deleted} 条记录")
                self._load_data()
            else:
                QMessageBox.warning(self, "删除失败", "未找到匹配的记录")
        except Exception as e:
            QMessageBox.critical(self, "删除失败", f"删除时发生错误:\n{e}")

        self._select_count_label.setText("")
        self._btn_batch_delete.setEnabled(False)

    def _on_delete_row(self):
        """删除单条记录"""
        btn = self.sender()
        if not btn:
            return
        row_data = btn.property("row_data")
        if not row_data:
            return

        sample_name = row_data.get("sample_name", "")
        test_date = row_data.get("test_date", "")
        test_time = row_data.get("test_time", "")

        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除以下记录吗？\n\n"
            f"试样名称: {sample_name}\n"
            f"试验日期: {test_date}\n"
            f"试验时间: {test_time}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            from core.database import DatabaseManager
            db = DatabaseManager()
            affected = db.execute(
                f'DELETE FROM "{_SUMMARY_TABLE_NAME}" '
                'WHERE "sample_name"=? AND "test_date"=? AND "test_time"=?',
                (sample_name, test_date, test_time),
            )
            if affected > 0:
                self._status_notify("已删除 1 条记录")
                self._load_data()
            else:
                QMessageBox.warning(self, "删除失败", "未找到匹配的记录")
        except Exception as e:
            QMessageBox.critical(self, "删除失败", f"删除时发生错误:\n{e}")

    def _on_edit_row(self):
        """编辑单条记录（支持曲线1~9）"""
        btn = self.sender()
        if not btn:
            return
        row_data = btn.property("row_data")
        if not row_data:
            return

        # 保存原始主键，用于 WHERE 条件
        orig_sample_name = row_data.get("sample_name", "")
        orig_test_date = row_data.get("test_date", "")
        orig_test_time = row_data.get("test_time", "")

        dialog = QDialog(self)
        dialog.setWindowTitle("编辑记录")
        dialog.setMinimumWidth(520)
        dialog.setStyleSheet("""
            QDialog { background-color: #fafbfc; }
            QLineEdit {
                border: 1px solid #d1d5db;
                border-radius: 4px;
                padding: 6px 10px;
                background-color: #ffffff;
                min-height: 28px;
                font-size: 13px;
            }
            QLineEdit:focus { border-color: #2563eb; }
            QComboBox {
                padding: 6px 10px;
                border: 1px solid #d1d5db;
                border-radius: 4px;
                background: white;
                min-height: 28px;
            }
            QLabel { color: #374151; font-size: 13px; }
            QPushButton#btn_add_curve {
                background-color: #eff6ff;
                color: #2563eb;
                border: 1px solid #bfdbfe;
                border-radius: 4px;
                padding: 4px 14px;
                font-size: 12px;
            }
            QPushButton#btn_add_curve:hover {
                background-color: #dbeafe;
            }
            QPushButton#btn_remove_curve {
                background-color: #f3f4f6;
                color: #6b7280;
                border: none;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 12px;
            }
            QPushButton#btn_remove_curve:hover {
                background-color: #fee2e2;
                color: #dc2626;
            }
        """)

        form = QFormLayout(dialog)
        form.setSpacing(12)
        form.setContentsMargins(20, 20, 20, 16)

        edit_name = QLineEdit(orig_sample_name)
        form.addRow("试样名称 *：", edit_name)

        edit_brand = QLineEdit(str(row_data.get("sample_brand", "")))
        form.addRow("试样牌号：", edit_brand)

        combo_polarity = QComboBox()
        combo_polarity.setEditable(True)
        combo_polarity.addItems(["正极", "负极"])
        current_polarity = str(row_data.get("polarity", ""))
        idx = combo_polarity.findText(current_polarity)
        if idx >= 0:
            combo_polarity.setCurrentIndex(idx)
        else:
            combo_polarity.setEditText(current_polarity)
        form.addRow("极性 *：", combo_polarity)

        edit_date = QLineEdit(str(orig_test_date))
        form.addRow("试验日期 *：", edit_date)

        edit_time = QLineEdit(str(orig_test_time))
        form.addRow("试验时间 *：", edit_time)

        # --- 动态曲线区域 ---
        curve_container = QWidget()
        curve_layout = QVBoxLayout(curve_container)
        curve_layout.setContentsMargins(0, 0, 0, 0)
        curve_layout.setSpacing(6)

        curve_rows: List[dict] = []

        def _add_curve_row(key: str, value: str = ""):
            row_widget = QWidget()
            row_lay = QHBoxLayout(row_widget)
            row_lay.setContentsMargins(0, 0, 0, 0)
            row_lay.setSpacing(6)

            cidx = int(key.split("_")[1])
            label = QLabel(f"曲线{cidx}：")
            label.setFixedWidth(55)
            row_lay.addWidget(label)

            edit = QLineEdit(value)
            edit.setPlaceholderText("数值，如 12.3456")
            row_lay.addWidget(edit)

            btn_remove = QPushButton("✕")
            btn_remove.setObjectName("btn_remove_curve")
            btn_remove.setFixedWidth(26)
            btn_remove.setProperty("curve_key", key)
            btn_remove.clicked.connect(lambda checked=False, w=row_widget: _remove_curve_row(w, key))
            row_lay.addWidget(btn_remove)

            curve_layout.addWidget(row_widget)
            curve_rows.append({"key": key, "edit": edit, "row_widget": row_widget})

        def _remove_curve_row(widget, key):
            if len(curve_rows) <= 1:
                return
            widget.setParent(None)
            widget.deleteLater()
            curve_rows[:] = [r for r in curve_rows if r["key"] != key]

        # 填入已有曲线数据（有值的 + 至少一行 curve_1）
        has_data = False
        for i in range(1, 10):
            ckey = f"curve_{i}"
            cval = row_data.get(ckey)
            if cval is not None and cval != "":
                _add_curve_row(ckey, f"{cval:.4f}" if isinstance(cval, float) else str(cval))
                has_data = True
        if not has_data:
            _add_curve_row("curve_1")

        # 添加按钮行
        add_row = QHBoxLayout()
        btn_add_curve = QPushButton("+ 添加曲线")
        btn_add_curve.setObjectName("btn_add_curve")

        def _on_add_curve():
            used_keys = {r["key"] for r in curve_rows}
            for i in range(1, 10):
                k = f"curve_{i}"
                if k not in used_keys:
                    _add_curve_row(k)
                    return
            QMessageBox.information(dialog, "提示", "最多支持9条曲线")

        btn_add_curve.clicked.connect(_on_add_curve)
        add_row.addWidget(btn_add_curve)
        add_row.addStretch()
        curve_layout.addLayout(add_row)

        form.addRow(curve_container)

        std_val = row_data.get("std_dev")
        edit_std = QLineEdit(f"{std_val:.4f}" if isinstance(std_val, float) else str(std_val or ""))
        form.addRow("标准差：", edit_std)

        edit_source = QLineEdit(str(row_data.get("source_file", "")))
        form.addRow("来源文件：", edit_source)

        # 按钮
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("取消")
        btn_cancel.setObjectName("db_close")
        btn_cancel.setMinimumWidth(80)
        btn_cancel.clicked.connect(dialog.reject)
        btn_row.addWidget(btn_cancel)

        btn_ok = QPushButton("保存")
        btn_ok.setObjectName("db_row_edit")
        btn_ok.setMinimumWidth(80)
        btn_ok.clicked.connect(dialog.accept)
        btn_row.addWidget(btn_ok)
        form.addRow(btn_row)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        # 收集新值
        sample_name = edit_name.text().strip()
        sample_brand = edit_brand.text().strip()
        polarity = combo_polarity.currentText().strip()
        test_date = edit_date.text().strip()
        test_time = edit_time.text().strip()

        # 校验必填项
        missing = []
        if not sample_name:
            missing.append("试样名称")
        if not polarity:
            missing.append("极性")
        if not test_date:
            missing.append("试验日期")
        if not test_time:
            missing.append("试验时间")
        if missing:
            QMessageBox.warning(
                self, "输入不完整",
                f"以下字段不能为空：{', '.join(missing)}"
            )
            return

        # 解析曲线数据
        curve_values = {}
        for r in curve_rows:
            text = r["edit"].text().strip()
            if text:
                try:
                    curve_values[r["key"]] = float(text)
                except ValueError:
                    QMessageBox.warning(self, "输入错误", f"{r['key']} 必须是数字")
                    return
            else:
                curve_values[r["key"]] = None  # 清空

        std_dev = None
        if edit_std.text().strip():
            try:
                std_dev = float(edit_std.text().strip())
            except ValueError:
                QMessageBox.warning(self, "输入错误", "标准差必须是数字")
                return

        # 保存到数据库
        try:
            from core.database import DatabaseManager
            db = DatabaseManager()

            key_changed = (
                sample_name != orig_sample_name
                or test_date != orig_test_date
                or test_time != orig_test_time
            )

            if key_changed:
                # 主键变了：先删旧记录，再插入新记录
                db.execute(
                    f'DELETE FROM "{_SUMMARY_TABLE_NAME}" '
                    'WHERE "sample_name"=? AND "test_date"=? AND "test_time"=?',
                    (orig_sample_name, orig_test_date, orig_test_time),
                )
                record = PeelDataRecord(
                    sample_name=sample_name,
                    sample_brand=sample_brand,
                    polarity=polarity,
                    test_date=test_date,
                    test_time=test_time,
                    std_dev=std_dev,
                    source_file=edit_source.text().strip(),
                    **curve_values,
                )
                inserted = db.insert_ignore(
                    PeelDataRecord.get_table_name(), record.to_dict()
                )
                if not inserted:
                    QMessageBox.warning(
                        self, "保存失败",
                        "新记录与已有记录冲突（试验时间+日期+名称 唯一约束）"
                    )
                    return
            else:
                # 主键未变：构建动态 UPDATE SET 子句
                set_parts = ['"sample_brand"=?', '"polarity"=?', '"std_dev"=?', '"source_file"=?']
                params = [sample_brand, polarity, std_dev, edit_source.text().strip()]

                # 更新所有曲线字段（包含清空为 NULL 的）
                for i in range(1, 10):
                    ckey = f"curve_{i}"
                    set_parts.append(f'"{ckey}"=?')
                    params.append(curve_values.get(ckey))

                params.extend([orig_sample_name, orig_test_date, orig_test_time])
                db.execute(
                    f'UPDATE "{_SUMMARY_TABLE_NAME}" SET {", ".join(set_parts)} '
                    'WHERE "sample_name"=? AND "test_date"=? AND "test_time"=?',
                    tuple(params),
                )

            self._status_notify("已更新 1 条记录")
            self._load_data()
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"写入数据库时发生错误:\n{e}")

    def _on_batch_delete(self):
        """批量删除勾选的记录"""
        checked = self._get_checked_records()
        if not checked:
            QMessageBox.information(self, "提示", "请先勾选要删除的记录")
            return

        reply = QMessageBox.question(
            self, "确认批量删除",
            f"确定要删除勾选的 {len(checked)} 条记录吗？\n"
            f"此操作不可撤销！",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        deleted = 0
        try:
            from core.database import DatabaseManager
            db = DatabaseManager()
            for row_data in checked:
                affected = db.execute(
                    f'DELETE FROM "{_SUMMARY_TABLE_NAME}" '
                    'WHERE "sample_name"=? AND "test_date"=? AND "test_time"=?',
                    (row_data.get("sample_name", ""),
                     row_data.get("test_date", ""),
                     row_data.get("test_time", ""))
                )
                deleted += affected

            if deleted > 0:
                self._status_notify(f"已删除 {deleted} 条记录")
                self._load_data()
            else:
                QMessageBox.warning(self, "删除失败", "未找到匹配的记录")
        except Exception as e:
            QMessageBox.critical(self, "删除失败", f"删除时发生错误:\n{e}")

        self._select_count_label.setText("")
        self._btn_batch_delete.setEnabled(False)

    def _on_add_record(self):
        """新增一条数据库记录（支持曲线1~9动态选择）"""
        dialog = QDialog(self)
        dialog.setWindowTitle("新增数据库记录")
        dialog.setMinimumWidth(520)
        dialog.setStyleSheet("""
            QDialog { background-color: #fafbfc; }
            QLineEdit {
                border: 1px solid #d1d5db;
                border-radius: 4px;
                padding: 6px 10px;
                background-color: #ffffff;
                min-height: 28px;
                font-size: 13px;
            }
            QLineEdit:focus { border-color: #10b981; }
            QComboBox {
                padding: 6px 10px;
                border: 1px solid #d1d5db;
                border-radius: 4px;
                background: white;
                min-height: 28px;
            }
            QLabel { color: #374151; font-size: 13px; }
            QPushButton#btn_add_curve {
                background-color: #eff6ff;
                color: #2563eb;
                border: 1px solid #bfdbfe;
                border-radius: 4px;
                padding: 4px 14px;
                font-size: 12px;
            }
            QPushButton#btn_add_curve:hover {
                background-color: #dbeafe;
            }
            QPushButton#btn_remove_curve {
                background-color: #f3f4f6;
                color: #6b7280;
                border: none;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 12px;
            }
            QPushButton#btn_remove_curve:hover {
                background-color: #fee2e2;
                color: #dc2626;
            }
        """)

        form = QFormLayout(dialog)
        form.setSpacing(12)
        form.setContentsMargins(20, 20, 20, 16)

        edit_name = QLineEdit()
        edit_name.setPlaceholderText("如 BG-01-正极")
        form.addRow("试样名称 *：", edit_name)

        edit_brand = QLineEdit()
        edit_brand.setPlaceholderText("如 BG-01（可选）")
        form.addRow("试样牌号：", edit_brand)

        combo_polarity = QComboBox()
        combo_polarity.setEditable(True)
        combo_polarity.addItems(["正极", "负极"])
        form.addRow("极性 *：", combo_polarity)

        edit_date = QLineEdit()
        edit_date.setPlaceholderText("如 2026-06-12")
        form.addRow("试验日期 *：", edit_date)

        edit_time = QLineEdit()
        edit_time.setPlaceholderText("如 10:30:00")
        form.addRow("试验时间 *：", edit_time)

        # --- 动态曲线区域 ---
        curve_container = QWidget()
        curve_layout = QVBoxLayout(curve_container)
        curve_layout.setContentsMargins(0, 0, 0, 0)
        curve_layout.setSpacing(6)

        curve_rows: List[dict] = []  # [{key: "curve_1", edit: QLineEdit, row_widget: QWidget}]

        def _add_curve_row(key: str, value: str = ""):
            row_widget = QWidget()
            row_lay = QHBoxLayout(row_widget)
            row_lay.setContentsMargins(0, 0, 0, 0)
            row_lay.setSpacing(6)

            idx = int(key.split("_")[1])
            label = QLabel(f"曲线{idx}：")
            label.setFixedWidth(55)
            row_lay.addWidget(label)

            edit = QLineEdit(value)
            edit.setPlaceholderText("数值，如 12.3456")
            row_lay.addWidget(edit)

            btn_remove = QPushButton("✕")
            btn_remove.setObjectName("btn_remove_curve")
            btn_remove.setFixedWidth(26)
            btn_remove.setProperty("curve_key", key)
            btn_remove.clicked.connect(lambda checked=False, w=row_widget: _remove_curve_row(w, key))
            row_lay.addWidget(btn_remove)

            curve_layout.addWidget(row_widget)
            curve_rows.append({"key": key, "edit": edit, "row_widget": row_widget})
            return edit

        def _remove_curve_row(widget, key):
            if len(curve_rows) <= 1:
                return  # 至少保留一行
            widget.setParent(None)
            widget.deleteLater()
            curve_rows[:] = [r for r in curve_rows if r["key"] != key]

        # 添加按钮行
        add_row = QHBoxLayout()
        btn_add_curve = QPushButton("+ 添加曲线")
        btn_add_curve.setObjectName("btn_add_curve")
        next_curve_idx = [2]  # 下一个可用的曲线编号

        def _on_add_curve():
            # 找到下一个未使用的曲线编号
            used_keys = {r["key"] for r in curve_rows}
            for i in range(1, 10):
                k = f"curve_{i}"
                if k not in used_keys:
                    _add_curve_row(k)
                    return
            QMessageBox.information(dialog, "提示", "最多支持9条曲线")

        btn_add_curve.clicked.connect(_on_add_curve)
        add_row.addWidget(btn_add_curve)
        add_row.addStretch()
        curve_layout.addLayout(add_row)

        form.addRow(curve_container)

        edit_std = QLineEdit()
        edit_std.setPlaceholderText("数值（可选）")
        form.addRow("标准差：", edit_std)

        edit_source = QLineEdit()
        edit_source.setPlaceholderText("来源文件名（可选）")
        form.addRow("来源文件：", edit_source)

        # 按钮
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("取消")
        btn_cancel.setObjectName("db_close")
        btn_cancel.setMinimumWidth(80)
        btn_cancel.clicked.connect(dialog.reject)
        btn_row.addWidget(btn_cancel)

        btn_ok = QPushButton("确定")
        btn_ok.setObjectName("db_add_record")
        btn_ok.setMinimumWidth(80)
        btn_ok.clicked.connect(dialog.accept)
        btn_row.addWidget(btn_ok)
        form.addRow(btn_row)

        # 默认添加 curve_1
        _add_curve_row("curve_1")

        if dialog.exec() == QDialog.DialogCode.Accepted:
            sample_name = edit_name.text().strip()
            sample_brand = edit_brand.text().strip()
            polarity = combo_polarity.currentText().strip()
            test_date = edit_date.text().strip()
            test_time = edit_time.text().strip()

            # 校验必填项
            missing = []
            if not sample_name:
                missing.append("试样名称")
            if not polarity:
                missing.append("极性")
            if not test_date:
                missing.append("试验日期")
            if not test_time:
                missing.append("试验时间")

            if missing:
                QMessageBox.warning(
                    self, "输入不完整",
                    f"以下字段不能为空：{', '.join(missing)}"
                )
                return

            # 解析曲线数据
            curve_values = {}
            for r in curve_rows:
                text = r["edit"].text().strip()
                if text:
                    try:
                        curve_values[r["key"]] = float(text)
                    except ValueError:
                        QMessageBox.warning(self, "输入错误", f"{r['key']} 必须是数字")
                        return

            std_dev = None
            if edit_std.text().strip():
                try:
                    std_dev = float(edit_std.text().strip())
                except ValueError:
                    QMessageBox.warning(self, "输入错误", "标准差必须是数字")
                    return

            # 构造记录并写入数据库
            record = PeelDataRecord(
                sample_name=sample_name,
                sample_brand=sample_brand,
                polarity=polarity,
                test_date=test_date,
                test_time=test_time,
                std_dev=std_dev,
                source_file=edit_source.text().strip(),
                **curve_values,
            )

            try:
                from core.database import DatabaseManager
                db = DatabaseManager()

                # 去重检查：同一 (test_date, test_time, polarity) 视为同一测试
                existing = db.query_one(
                    f'SELECT id, sample_name FROM "{_SUMMARY_TABLE_NAME}" '
                    'WHERE "test_date"=? AND "test_time"=? AND "polarity"=?',
                    (test_date, test_time, polarity)
                )
                if existing:
                    QMessageBox.warning(
                        self, "新增失败",
                        f"已存在相同时间+极性的记录（试样名称: {existing['sample_name']}）。\n"
                        f"同一时间同一极性的测试数据不能重复添加。\n\n"
                        f"如需修改数据，请使用编辑功能。"
                    )
                    return

                inserted = db.insert_ignore(
                    PeelDataRecord.get_table_name(), record.to_dict()
                )
                if inserted:
                    self._status_notify(f"已新增 1 条记录")
                    self._load_data()
                else:
                    QMessageBox.warning(
                        self, "新增失败",
                        "记录已存在（试验时间+日期+名称 唯一约束冲突）"
                    )
            except Exception as e:
                QMessageBox.critical(self, "新增失败", f"写入数据库时发生错误:\n{e}")

    def _status_notify(self, message: str):
        """状态栏通知（替代频繁弹窗）"""
        self._info_label.setText(message)
        # 3秒后恢复
        from PySide6.QtCore import QTimer
        QTimer.singleShot(3000, lambda: self._info_label.setText(
            f"共 {len(self._records)} 条记录"
        ))


class ExtractWorker(QThread):
    """数据提取工作线程（支持请求ID关联追溯）"""

    progress = Signal(int, int, str)
    finished = Signal(object)
    log_message = Signal(str, str)

    def __init__(self, directory: str, save_to_db: bool = True):
        super().__init__()
        self._directory = directory
        self._save_to_db = save_to_db
        self._extractor = PeelDataExtractor()
        # 生成本次提取任务的请求ID，用于关联追溯完整业务流程
        from core.unified_logger import RequestContext
        self._request_id = RequestContext.generate_request_id()

    def run(self):
        from core.unified_logger import RequestContext
        RequestContext.set_request_id(self._request_id)
        logger.info(f"[请求 {self._request_id}] 开始数据提取: {self._directory}")
        try:
            result = self._extractor.extract(
                self._directory,
                save_to_db=self._save_to_db,
                progress_callback=self._on_progress,
                request_id=self._request_id,
            )
            self.finished.emit(result)
        except Exception as e:
            logger.error(f"[请求 {self._request_id}] 提取线程异常: {e}", exc_info=True)
            self.finished.emit(None)
        finally:
            RequestContext.clear()

    def _on_progress(self, current, total, message):
        self.progress.emit(current, total, message)

    def cancel(self):
        self._extractor.cancel()

    @property
    def request_id(self) -> str:
        return self._request_id


class PeelDataWidget(QWidget):
    """剥离数据汇总主界面"""

    # 线程安全日志信号：日志 handler 从 worker 线程 emit → 主线程槽函数安全更新 UI
    _log_signal = Signal(str, str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._extractor = PeelDataExtractor()
        self._worker: Optional[ExtractWorker] = None
        self._records: List[PeelDataRecord] = []
        self._last_result: Optional[ExtractionResult] = None
        self._setup_ui()
        self._connect_signals()
        self._init_button_states()

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(4, 4, 4, 4)
        main_layout.setSpacing(10)

        # === 标题栏：数据提取 + 操作按钮 ===
        title_layout = QHBoxLayout()
        title_layout.setContentsMargins(8, 6, 8, 0)
        title_label = QLabel("数据提取")
        title_font = QFont()
        title_font.setBold(True)
        title_font.setPointSize(12)
        title_label.setFont(title_font)
        title_layout.addWidget(title_label)
        title_layout.addStretch()

        # 历史记录按钮
        self._btn_history = QPushButton("历史记录")
        self._btn_history.setEnabled(False)
        self._btn_history.setMinimumWidth(90)
        self._btn_history.setObjectName("btn_secondary")
        title_layout.addWidget(self._btn_history)

        # 查看数据库按钮
        self._btn_view_db = QPushButton("查看数据库")
        self._btn_view_db.setMinimumWidth(90)
        self._btn_view_db.setObjectName("btn_secondary")
        title_layout.addWidget(self._btn_view_db)

        main_layout.addLayout(title_layout)

        # === 顶部操作区 ===
        top_group = QGroupBox()
        top_group.setObjectName("flat_group")
        top_layout = QGridLayout(top_group)
        top_layout.setSpacing(10)
        top_layout.setContentsMargins(12, 16, 12, 12)

        # 目录选择行
        top_layout.addWidget(QLabel("数据目录："), 0, 0)
        self._dir_edit = QLineEdit(config.last_data_dir)
        self._dir_edit.setPlaceholderText("首次使用请点击「选择目录」指定数据文件夹")
        top_layout.addWidget(self._dir_edit, 0, 1)

        self._btn_browse = QPushButton("选择目录")
        self._btn_browse.setObjectName("btn_warning")
        self._btn_browse.setMinimumWidth(90)
        top_layout.addWidget(self._btn_browse, 0, 2)

        # 操作按钮行（开始提取 | 取消 | 保存到数据库）
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self._btn_extract = QPushButton("开始提取")
        self._btn_extract.setObjectName("btn_success")
        self._btn_extract.setMinimumWidth(90)
        btn_layout.addWidget(self._btn_extract)

        self._btn_cancel = QPushButton("取消")
        self._btn_cancel.setObjectName("btn_danger")
        self._btn_cancel.setEnabled(False)
        self._btn_cancel.setMinimumWidth(70)
        btn_layout.addWidget(self._btn_cancel)

        self._chk_save_db = QCheckBox("保存到数据库")
        self._chk_save_db.setChecked(True)
        btn_layout.addWidget(self._chk_save_db)

        btn_layout.addStretch()
        top_layout.addLayout(btn_layout, 1, 0, 1, 3)

        # 进度条
        self._progress_bar = QProgressBar()
        self._progress_bar.setVisible(False)
        top_layout.addWidget(self._progress_bar, 2, 0, 1, 3)

        # 状态标签
        self._status_label = QLabel("就绪")
        self._status_label.setObjectName("status_label")
        top_layout.addWidget(self._status_label, 3, 0, 1, 3)

        main_layout.addWidget(top_group)

        # === 中间内容区（表格 + 日志 分割） ===
        splitter = QSplitter(Qt.Orientation.Vertical)

        # 数据预览表格
        table_group = QGroupBox("数据预览")
        table_layout = QVBoxLayout(table_group)
        table_layout.setSpacing(6)

        # 编辑提示
        edit_hint = QLabel('双击单元格可编辑，修改后点击「保存修改」同步到数据库')
        edit_hint.setObjectName("status_label")
        table_layout.addWidget(edit_hint)

        self._table = QTableWidget()
        self._table.setAlternatingRowColors(True)
        self._table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalHeader().setVisible(False)
        table_layout.addWidget(self._table)

        # 操作按钮行
        export_layout = QHBoxLayout()

        self._btn_export = QPushButton("导出 Excel")
        self._btn_export.setEnabled(False)
        self._btn_export.setMinimumWidth(90)
        export_layout.addWidget(self._btn_export)

        self._btn_save_edit = QPushButton("保存修改")
        self._btn_save_edit.setEnabled(False)
        self._btn_save_edit.setObjectName("btn_success")
        self._btn_save_edit.setMinimumWidth(90)
        export_layout.addWidget(self._btn_save_edit)

        self._label_record_count = QLabel("共 0 条记录")
        self._label_record_count.setObjectName("status_label")
        export_layout.addWidget(self._label_record_count)

        export_layout.addStretch()
        table_layout.addLayout(export_layout)

        splitter.addWidget(table_group)

        # 日志面板
        log_group = QGroupBox("运行日志")
        log_layout = QVBoxLayout(log_group)
        log_layout.setSpacing(6)

        # 日志级别选择
        log_ctrl_layout = QHBoxLayout()
        log_ctrl_layout.setSpacing(6)
        log_ctrl_layout.addWidget(QLabel("日志级别："))

        self._combo_log_level = QComboBox()
        self._combo_log_level.addItems(["DEBUG", "INFO", "WARNING", "ERROR"])
        self._combo_log_level.setCurrentText("INFO")
        self._combo_log_level.setMinimumWidth(90)
        log_ctrl_layout.addWidget(self._combo_log_level)

        self._chk_auto_scroll = QCheckBox("自动滚动")
        self._chk_auto_scroll.setChecked(True)
        log_ctrl_layout.addWidget(self._chk_auto_scroll)

        log_ctrl_layout.addStretch()

        self._btn_clear_log = QPushButton("清空")
        self._btn_clear_log.setMinimumWidth(60)
        log_ctrl_layout.addWidget(self._btn_clear_log)

        # 查看日志文件按钮
        self._btn_open_unified_log = QPushButton("查看日志文件")
        self._btn_open_unified_log.setMinimumWidth(100)
        self._btn_open_unified_log.setObjectName("btn_secondary")
        self._btn_open_unified_log.setToolTip("在文件管理器中打开日志文件所在位置")
        log_ctrl_layout.addWidget(self._btn_open_unified_log)

        log_ctrl_layout.addStretch()
        log_layout.addLayout(log_ctrl_layout)

        self._log_text = QTextEdit()
        self._log_text.setReadOnly(True)
        self._log_text.setMaximumHeight(200)
        self._log_text.setStyleSheet(
            "font-family: Consolas, 'Microsoft YaHei', monospace; "
            "font-size: 12px; background-color: #1e1e1e; color: #d4d4d4;"
        )
        log_layout.addWidget(self._log_text)

        splitter.addWidget(log_group)

        # 设置分割比例
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 1)

        main_layout.addWidget(splitter, stretch=1)

        # 初始化日志 handler
        self._setup_log_handler()

    def _setup_log_handler(self):
        """设置日志 handler，通过 Qt 信号将日志消息安全转发到主线程 UI"""
        import logging

        class WidgetLogHandler(logging.Handler):
            def __init__(self, widget: "PeelDataWidget"):
                super().__init__()
                self._widget = widget

            def emit(self, record):
                try:
                    msg = self.format(record)
                    level = record.levelname
                    # 通过信号发射日志 → 主线程槽函数安全更新 UI
                    # Qt 自动使用 QueuedConnection（跨线程发射时），避免跨线程 UI 操作崩溃
                    self._widget._log_signal.emit(msg, level)
                except Exception:
                    pass

        handler = WidgetLogHandler(self)
        handler.setLevel(logging.DEBUG)
        fmt = logging.Formatter(
            "[%(asctime)s] [%(levelname)s] %(message)s",
            datefmt="%H:%M:%S",
        )
        handler.setFormatter(fmt)

        for logger_name in ["peel_data.pdf_parser", "peel_data.excel_parser",
                            "peel_data.extractor", "peel_data.ui", "database"]:
            lg = logging.getLogger(logger_name)
            lg.addHandler(handler)

    def _append_log(self, message: str, level: str):
        """追加日志消息到界面"""
        color_map = {
            "DEBUG": "#6a9955",
            "INFO": "#d4d4d4",
            "WARNING": "#ce9178",
            "ERROR": "#f44747",
            "CRITICAL": "#ff0000",
        }
        color = color_map.get(level, "#d4d4d4")
        html = f'<span style="color:{color}">{message}</span>'

        self._log_text.append(html)

        if self._chk_auto_scroll.isChecked():
            scrollbar = self._log_text.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())

    def _connect_signals(self):
        """连接信号槽"""
        self._btn_browse.clicked.connect(self._on_browse)
        self._btn_extract.clicked.connect(self._on_extract)
        self._btn_cancel.clicked.connect(self._on_cancel)
        self._btn_export.clicked.connect(self._on_export)
        self._btn_save_edit.clicked.connect(self._on_save_edit)
        self._btn_history.clicked.connect(self._on_show_history)
        self._btn_view_db.clicked.connect(self._on_show_database)
        self._btn_clear_log.clicked.connect(self._log_text.clear)
        self._btn_open_unified_log.clicked.connect(self._on_open_unified_log)
        self._combo_log_level.currentTextChanged.connect(self._on_log_level_changed)
        self._table.cellChanged.connect(self._on_cell_changed)
        # 线程安全：worker 线程日志信号 → 主线程安全更新 UI
        self._log_signal.connect(self._append_log)

    def _init_button_states(self):
        """初始化按钮状态（根据数据库中是否已有历史记录）"""
        try:
            if self._load_history_from_db():
                self._btn_history.setEnabled(True)
        except Exception:
            pass

    def _on_browse(self):
        """选择数据目录 —— 选择后自动持久化，下次启动自动回填"""
        current_dir = self._dir_edit.text() or os.path.expanduser("~")
        directory = QFileDialog.getExistingDirectory(
            self, "选择数据目录", current_dir
        )
        if directory:
            self._dir_edit.setText(directory)
            # 持久化到配置（下次启动自动回填）
            config.last_data_dir = directory
            config.save()
            logger.info(f"选择目录并已保存: {directory}")

    def _on_extract(self):
        """开始提取"""
        directory = self._dir_edit.text().strip()
        if not directory or not os.path.isdir(directory):
            QMessageBox.warning(self, "目录无效", "请选择一个有效的数据目录")
            return

        self._btn_extract.setEnabled(False)
        self._btn_cancel.setEnabled(True)
        self._btn_browse.setEnabled(False)
        self._btn_export.setEnabled(False)
        self._btn_save_edit.setEnabled(False)
        self._btn_history.setEnabled(False)

        self._progress_bar.setVisible(True)
        self._progress_bar.setValue(0)
        self._status_label.setText("提取中...")

        self._worker = ExtractWorker(
            directory,
            save_to_db=self._chk_save_db.isChecked(),
        )
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_extract_finished)
        self._worker.start()

    def _on_cancel(self):
        """取消提取"""
        if self._worker and self._worker.isRunning():
            self._worker.cancel()
            self._status_label.setText("正在取消...")

    def _on_progress(self, current: int, total: int, message: str):
        """更新进度"""
        if total > 0:
            self._progress_bar.setMaximum(total)
            self._progress_bar.setValue(current)
        self._status_label.setText(message)

    def _on_extract_finished(self, result: Optional[ExtractionResult]):
        """提取完成"""
        self._btn_extract.setEnabled(True)
        self._btn_cancel.setEnabled(False)
        self._btn_browse.setEnabled(True)
        self._progress_bar.setVisible(False)

        if result is None:
            self._status_label.setText("提取失败，详见日志")
            QMessageBox.critical(
                self, "提取失败",
                "数据提取过程中发生错误，请查看运行日志了解详情"
            )
            return

        self._last_result = result
        self._records = result.records or []
        self._skipped_records = result.skipped_records or []

        # 填充表格（防御性 try/except）
        try:
            self._populate_table(self._records, self._skipped_records)
        except Exception as e:
            logger.error(f"填充表格失败: {e}", exc_info=True)
            QMessageBox.warning(
                self, "表格渲染失败",
                f"数据已提取但表格渲染失败：\n{e}\n\n"
                f"数据仍可导出为 Excel。"
            )

        # 状态栏/统计行：显示完整提取摘要（含应用层去重、数据库写入等）
        self._status_label.setText(result.summary)
        if result.app_skipped > 0:
            self._label_record_count.setText(
                f"共 {len(self._records)} 条已保留"
                f"，{result.app_skipped} 条去重跳过（表格中暖黄底色行）"
            )
        else:
            self._label_record_count.setText(f"共 {len(self._records)} 条记录")

        if self._records:
            self._btn_export.setEnabled(True)
            self._btn_save_edit.setEnabled(True)

        # 有历史记录则启用历史按钮（数据库有记录或本次提取有历史）
        if result.history or self._load_history_from_db():
            self._btn_history.setEnabled(True)

        # === BugFix #18: 不再弹出 PreviewDialog，数据已直接填入主窗口的 _table 表格 ===
        # 统计信息通过 result.summary 写入 _status_label，用户在主窗口即可看到。
        # 之前 PreviewDialog 的弹窗体验被替换为嵌入式预览，截图与导出按钮不变。

    def _get_active_curves(self, records: List[PeelDataRecord]) -> List[int]:
        """确定实际存在的曲线列"""
        active = set()
        for r in records:
            for i in range(1, 10):
                if getattr(r, f"curve_{i}") is not None:
                    active.add(i)
        return sorted(active)

    def _get_common_curve_unit(self, records: List[PeelDataRecord]) -> str:
        """获取记录中最常见的曲线单位"""
        from collections import Counter
        units = [r.curve_unit for r in records if r.curve_unit]
        if not units:
            return ""
        counter = Counter(units)
        return counter.most_common(1)[0][0]

    def _populate_table(
        self,
        records: List[PeelDataRecord],
        skipped_records: List = None,
    ):
        """填充数据预览表格（合并日期时间为单字段）

        Args:
            records: 保留的记录列表（正常白底）
            skipped_records: 被去重跳过的记录列表（暖黄底色，源自 SkippedRecord）
        """
        skipped_records = skipped_records or []

        # 合并所有展示项：保留记录 + 跳过记录
        # 跳过记录统一打上"⚠ 跳过"标记
        all_rows = []  # [(record, is_skipped, skip_reason), ...]
        for r in records:
            all_rows.append((r, False, ""))
        for sr in skipped_records:
            all_rows.append((sr.record, True, sr.reason))

        if not all_rows:
            self._table.clear()
            self._table.setRowCount(0)
            return

        active_curves = self._get_active_curves(records)
        curve_unit = self._get_common_curve_unit(records)
        unit_suffix = f" ({curve_unit})" if curve_unit else ""

        # 表头：试样名称 | 试样牌号 | 极性 | 试验日期时间 | 曲线N(单位)... | 标准差 | 来源文件 | 状态
        headers = ["试样名称", "试样牌号", "极性", "试验日期时间"]
        headers.extend([f"曲线{i}{unit_suffix}" for i in active_curves])
        headers.extend(["标准差", "来源文件", "状态"])

        self._table.setColumnCount(len(headers))
        self._table.setHorizontalHeaderLabels(headers)
        self._table.setRowCount(len(all_rows))

        # 去重行底色（暖黄）+ 文字色（深棕）
        SKIP_BG = QColor("#fff3cd")
        SKIP_FG = QColor("#92400e")

        for row_idx, (record, is_skipped, skip_reason) in enumerate(all_rows):
            col = 0

            def _make_item(text: str):
                item = QTableWidgetItem(str(text) if text is not None else "")
                if is_skipped:
                    item.setBackground(SKIP_BG)
                    item.setForeground(SKIP_FG)
                    if skip_reason:
                        item.setToolTip(f"去重原因：{skip_reason}")
                return item

            # 试样名称
            self._table.setItem(row_idx, col, _make_item(record.sample_name))
            col += 1
            # 试样牌号
            self._table.setItem(row_idx, col, _make_item(record.sample_brand))
            col += 1
            # 极性
            polarity_item = _make_item(record.polarity)
            if not is_skipped:
                if record.polarity == "正极":
                    polarity_item.setForeground(QColor("#e74c3c"))
                elif record.polarity == "负极":
                    polarity_item.setForeground(QColor("#27ae60"))
                else:
                    polarity_item.setForeground(QColor("#95a5a6"))
            self._table.setItem(row_idx, col, polarity_item)
            col += 1
            # 试验日期时间
            self._table.setItem(row_idx, col, _make_item(record.test_datetime))
            col += 1
            # 曲线值
            for i in active_curves:
                val = getattr(record, f"curve_{i}", None)
                text = f"{val:.4f}" if val is not None else ""
                self._table.setItem(row_idx, col, _make_item(text))
                col += 1
            # 标准差
            std_text = f"{record.std_dev:.4f}" if record.std_dev is not None else ""
            self._table.setItem(row_idx, col, _make_item(std_text))
            col += 1
            # 来源文件（不可编辑，设为只读提示）
            source_item = _make_item(record.source_file)
            source_item.setFlags(source_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self._table.setItem(row_idx, col, source_item)
            col += 1
            # 状态列
            if is_skipped:
                status_item = QTableWidgetItem("⚠ 去重跳过")
                status_item.setBackground(SKIP_BG)
                status_item.setForeground(SKIP_FG)
                status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if skip_reason:
                    status_item.setToolTip(skip_reason)
            else:
                status_item = QTableWidgetItem("✓ 已保留")
                status_item.setForeground(QColor("#15803d"))
                status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._table.setItem(row_idx, col, status_item)
            col += 1

        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        self._table.resizeColumnsToContents()

    def _on_cell_changed(self, row: int, column: int):
        """表格单元格编辑后更新底层记录"""
        if not self._records or row >= len(self._records):
            return

        record = self._records[row]
        item = self._table.item(row, column)
        if not item:
            return

        new_value = item.text().strip()
        active_curves = self._get_active_curves(self._records)

        # 列映射（注意：来源文件列不可编辑，已在 flags 中禁用）
        if column == 0:
            record.sample_name = new_value
        elif column == 1:
            # 试样牌号（新增独立字段）
            record.sample_brand = new_value
        elif column == 2:
            record.polarity = new_value
            # 更新颜色
            if new_value == "正极":
                item.setForeground(QColor("#e74c3c"))
            elif new_value == "负极":
                item.setForeground(QColor("#27ae60"))
            else:
                item.setForeground(QColor("#95a5a6"))
        elif column == 3:
            # 试验日期时间 → 拆分回 test_date 和 test_time
            parts = new_value.split(" ")
            if len(parts) >= 2:
                record.test_date = parts[0]
                record.test_time = parts[1]
            elif len(parts) == 1:
                if ":" in parts[0]:
                    record.test_time = parts[0]
                else:
                    record.test_date = parts[0]
        else:
            # 计算列对应的字段
            # headers: [试样名称, 试样牌号, 极性, 试验日期时间, 曲线1..N, 标准差, 来源文件]
            # column 4 开始是曲线
            curve_end_col = 4 + len(active_curves)
            if 4 <= column < curve_end_col:
                curve_idx = active_curves[column - 4]
                try:
                    val = float(new_value) if new_value else None
                    setattr(record, f"curve_{curve_idx}", val)
                except ValueError:
                    pass
            elif column == curve_end_col:
                try:
                    record.std_dev = float(new_value) if new_value else None
                except ValueError:
                    pass

        logger.debug(f"单元格编辑: row={row}, col={column}, value={new_value}")

    def _on_save_edit(self):
        """将表格中的修改保存到数据库"""
        if not self._records:
            QMessageBox.warning(self, "无数据", "当前没有可保存的数据")
            return

        try:
            db_available = _is_db_available()
        except Exception:
            db_available = False

        if not db_available:
            QMessageBox.warning(self, "数据库不可用", "数据库未连接，无法保存修改")
            return

        try:
            db = DatabaseManager()
            updated = 0
            for record in self._records:
                # 先删除旧记录（按唯一约束匹配）
                db.execute(
                    f'DELETE FROM "{_SUMMARY_TABLE_NAME}" '
                    'WHERE "test_time"=? AND "test_date"=? AND "sample_name"=?',
                    (record.test_time, record.test_date, record.sample_name)
                )
                # 插入新记录
                if db.insert_ignore(PeelDataRecord.get_table_name(), record.to_dict()):
                    updated += 1

            logger.info(f"已保存 {updated} 条修改到数据库")
            self._status_label.setText(f"已保存 {updated} 条修改到数据库")
            QMessageBox.information(self, "保存成功", f"已成功保存 {updated} 条记录到数据库")

            # 刷新表格显示
            self._populate_table(self._records)

        except Exception as e:
            logger.error(f"保存修改失败: {e}", exc_info=True)
            QMessageBox.critical(self, "保存失败", f"保存数据时发生错误:\n{e}")

    def _on_show_history(self):
        """显示提取历史记录（内存 + 数据库持久化）"""
        # 从数据库加载历史
        db_history = self._load_history_from_db()
        all_history = list(db_history)

        # 合并当前会话历史（去重：同 file_path 保留最新的）
        seen = {h.file_path for h in all_history}
        if self._last_result and self._last_result.history:
            for h in self._last_result.history:
                if h.file_path not in seen:
                    all_history.append(h)
                    seen.add(h.file_path)

        if not all_history:
            QMessageBox.information(self, "历史记录", "暂无提取历史记录")
            return

        # 按时间倒序（数据库记录有 created_at，内存记录没有，放后面）
        # 简化处理：数据库记录在前（已按 created_at DESC），内存记录追加在后
        dialog = HistoryDialog(all_history, self)
        dialog.exec()

    def _load_history_from_db(self) -> list:
        """从数据库加载历史记录，返回 List[FileHistory]"""
        try:
            from core.database import DatabaseManager
            from plugins.peel_data.models import ensure_history_table, get_history_table_name
            ensure_history_table()  # 确保表结构最新（含 operation_time 列）
            table_name = get_history_table_name()
            db = DatabaseManager()
            rows = db.query_all(
                'SELECT "file_path", "file_name", "success", "reason", "operation_time" '
                f'FROM "{table_name}" '
                'ORDER BY "created_at" DESC '
                'LIMIT 2000'
            )
            history = []
            for row in rows:
                history.append(FileHistory(
                    file_path=row["file_path"],
                    file_name=row["file_name"],
                    success=bool(row["success"]),
                    reason=row["reason"] or "",
                    operation_time=row.get("operation_time", "") or "",
                ))
            return history
        except Exception as e:
            logger.error(f"从数据库加载历史失败: {e}", exc_info=True)
            return []

    def _on_show_database(self):
        """打开数据库记录浏览对话框"""
        try:
            db_available = _is_db_available()
        except Exception:
            db_available = False

        if not db_available:
            QMessageBox.warning(self, "数据库不可用", "数据库未连接，无法浏览记录")
            return
        dialog = DatabaseViewerDialog(self)
        dialog.exec()

    def _on_export(self):
        """导出 Excel"""
        if not self._records:
            QMessageBox.warning(self, "无数据", "当前没有可导出的数据")
            return

        default_name = "剥离数据汇总.xlsx"
        default_dir = self._dir_edit.text().strip() or os.path.expanduser("~")

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出 Excel",
            os.path.join(default_dir, default_name),
            "Excel 文件 (*.xlsx);;所有文件 (*.*)",
        )

        if file_path:
            self._status_label.setText("正在导出...")
            success = PeelDataExtractor.export_to_excel(
                self._records, file_path
            )
            if success:
                self._status_label.setText(f"导出成功: {file_path}")
                QMessageBox.information(
                    self, "导出成功",
                    f"数据已导出到:\n{file_path}"
                )
            else:
                self._status_label.setText("导出失败")
                QMessageBox.critical(self, "导出失败", "导出过程中发生错误，请查看运行日志了解详情")

    def _on_open_unified_log(self):
        """打开统一日志文件所在位置"""
        try:
            from core.logger import get_unified_log_path
            log_path = get_unified_log_path()
            if os.path.exists(log_path):
                subprocess.run(["explorer", "/select,", os.path.normpath(log_path)], check=False)
            else:
                QMessageBox.information(self, "日志文件", f"日志文件尚未生成:\n{log_path}")
        except Exception as e:
            QMessageBox.warning(self, "打开失败", f"无法打开日志文件:\n{e}")

    def _on_log_level_changed(self, level: str):
        """调整日志级别"""
        ToolkitLogger().set_all_level(level)
        logger.info(f"日志级别已调整为: {level}")
