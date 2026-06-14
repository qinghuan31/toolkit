# -*- coding: utf-8 -*-
"""
数据提取器 —— 协调 PDF/Excel 解析器，提供统一的提取接口
负责文件遍历、解析调度、数据存储和统计
"""

import os
import datetime as _dt
from typing import List, Tuple, Optional
from dataclasses import dataclass, field

from core.logger import get_logger
from core.database import DatabaseManager
from plugins.peel_data.models import PeelDataRecord, ensure_table, ensure_history_table, get_history_table_name
from plugins.peel_data.pdf_parser import PDFParser
from plugins.peel_data.excel_parser import ExcelParser

logger = get_logger("peel_data.extractor")


@dataclass
class FileHistory:
    """单条文件提取历史记录"""
    file_path: str = ""
    file_name: str = ""
    success: bool = False
    reason: str = ""          # 成功时说明提取到了哪些字段；失败时说明原因
    operation_time: str = ""  # 提取操作时间戳（格式 yyyy-MM-dd HH:mm:ss）


@dataclass
class SkippedRecord:
    """被去重跳过的记录（含跳过原因）"""
    record: PeelDataRecord
    reason: str  # "批次内重复" / "数据库已存在"
    matched_existing_id: int = None  # 跨批次跳过时，数据库中已有记录的 id


@dataclass
class ExtractionResult:
    """提取结果统计"""
    total_files: int = 0
    pdf_count: int = 0
    excel_count: int = 0
    csv_count: int = 0
    success_count: int = 0
    fail_count: int = 0
    db_inserted: int = 0
    db_skipped: int = 0
    app_skipped: int = 0            # 应用层去重跳过的记录数（批次内 + 跨批次）
    db_unavailable: bool = False
    positive_count: int = 0
    negative_count: int = 0
    s_complete: int = 0
    s_partial: int = 0
    records: List[PeelDataRecord] = None
    skipped_records: List[SkippedRecord] = None  # 被去重跳过的记录（UI 可展示）
    history: List[FileHistory] = None

    def __post_init__(self):
        if self.records is None:
            self.records = []
        if self.skipped_records is None:
            self.skipped_records = []
        if self.history is None:
            self.history = []

    @property
    def summary(self) -> str:
        db_info = ""
        if self.db_unavailable:
            db_info = "，数据库不可用（数据仅在内存中，可导出）"
        elif self.db_inserted > 0 or self.db_skipped > 0 or self.app_skipped > 0:
            db_info = (
                f"，数据库插入 {self.db_inserted} 条，"
                f"跳过重复 {self.db_skipped} 条"
            )
            if self.app_skipped > 0:
                db_info += f"，应用层去重 {self.app_skipped} 条"
        return (
            f"共扫描 {self.total_files} 个文件"
            f"（PDF {self.pdf_count}, Excel {self.excel_count}, CSV {self.csv_count}），"
            f"成功提取 {self.success_count} 条，"
            f"失败 {self.fail_count} 条"
            f"{db_info}"
        )


def _is_db_available() -> bool:
    """检查数据库是否可用（延迟检测）"""
    try:
        db = DatabaseManager()
        return db.is_available
    except Exception:
        return False


class PeelDataExtractor:
    """剥离数据提取器"""

    SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".xlsm", ".xlsb", ".csv"}

    def __init__(self):
        self._pdf_parser = PDFParser()
        self._excel_parser = ExcelParser()
        self._cancelled = False

    def scan_directory(self, directory: str) -> Tuple[List[str], List[str], List[str]]:
        """扫描目录，返回 (pdf_files, excel_files, csv_files)"""
        pdf_files = []
        excel_files = []
        csv_files = []

        if not os.path.isdir(directory):
            logger.error(f"目录不存在: {directory}")
            return pdf_files, excel_files, csv_files

        logger.info(f"开始扫描目录: {directory}")

        for root, dirs, files in os.walk(directory):
            for fname in files:
                ext = os.path.splitext(fname)[1].lower()
                fpath = os.path.join(root, fname)

                if ext == ".pdf":
                    pdf_files.append(fpath)
                elif ext in (".xlsx", ".xls", ".xlsm", ".xlsb"):
                    excel_files.append(fpath)
                elif ext == ".csv":
                    csv_files.append(fpath)

        logger.info(
            f"扫描完成: 发现 {len(pdf_files)} 个PDF, "
            f"{len(excel_files)} 个Excel, {len(csv_files)} 个CSV"
        )
        return pdf_files, excel_files, csv_files

    def extract(
        self,
        directory: str,
        save_to_db: bool = True,
        progress_callback=None,
        request_id: str = "",
        file_type: str = "auto",
    ) -> ExtractionResult:
        """
        执行数据提取主流程

        Args:
            directory: 源文件目录
            save_to_db: 是否保存到数据库
            progress_callback: 进度回调函数
            request_id: 请求 ID（用于历史记录）
            file_type: 文件类型过滤，"auto"|"pdf"|"excel"

        Returns:
            ExtractionResult: 提取结果统计
        """
        self._cancelled = False
        # 捕获本次提取操作的统一时间戳（所有文件共享同一批次的提取时间）
        op_time = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        result = ExtractionResult()

        # 确保数据表存在
        if save_to_db:
            if _is_db_available():
                ensure_table()
            else:
                logger.warning("数据库不可用，将跳过数据库写入（数据仍可预览和导出）")
                save_to_db = False
                result.db_unavailable = True

        # 扫描文件
        pdf_files, excel_files, csv_files = self.scan_directory(directory)

        # 根据 file_type 过滤文件
        if file_type == "pdf":
            excel_files = []
            csv_files = []
        elif file_type == "excel":
            pdf_files = []

        result.pdf_count = len(pdf_files)
        result.excel_count = len(excel_files)
        result.csv_count = len(csv_files)
        result.total_files = result.pdf_count + result.excel_count + result.csv_count

        if result.total_files == 0:
            logger.warning("未发现任何可处理文件")
            return result

        all_records: List[PeelDataRecord] = []
        processed = 0

        # 解析 PDF
        for fpath in pdf_files:
            if self._cancelled:
                break
            processed += 1
            if progress_callback:
                progress_callback(
                    processed, result.total_files,
                    f"正在解析: {os.path.basename(fpath)}"
                )

            record = self._pdf_parser.parse(fpath)
            if record:
                all_records.append(record)
                result.success_count += 1
                curve_count = sum(1 for i in range(1, 10) if getattr(record, f"curve_{i}") is not None)
                result.s_complete += 1 if curve_count >= 4 else (1 if curve_count > 0 else 0)
                if curve_count > 0 and curve_count < 4:
                    result.s_partial += 1
                if record.polarity == "正极":
                    result.positive_count += 1
                elif record.polarity == "负极":
                    result.negative_count += 1
                reason = f"试样={record.sample_name}, 极性={record.polarity}, 曲线{curve_count}条"
                result.history.append(FileHistory(
                    file_path=fpath,
                    file_name=os.path.basename(fpath),
                    success=True,
                    reason=reason,
                    operation_time=op_time,
                ))
                logger.debug(f"PDF 提取成功: {record.sample_name} 曲线数={curve_count}")
            else:
                result.fail_count += 1
                result.history.append(FileHistory(
                    file_path=fpath,
                    file_name=os.path.basename(fpath),
                    success=False,
                    reason="未提取到有效数据（缺少试样名称或剥离强度曲线）",
                    operation_time=op_time,
                ))
                logger.debug(f"PDF 提取失败: {os.path.basename(fpath)}")

        # 解析 Excel
        for fpath in excel_files:
            if self._cancelled:
                break
            processed += 1
            if progress_callback:
                progress_callback(
                    processed, result.total_files,
                    f"正在解析: {os.path.basename(fpath)}"
                )

            record = self._excel_parser.parse(fpath)
            if record:
                all_records.append(record)
                result.success_count += 1
                curve_count = sum(1 for i in range(1, 10) if getattr(record, f"curve_{i}") is not None)
                result.s_complete += 1 if curve_count >= 4 else (1 if curve_count > 0 else 0)
                if curve_count > 0 and curve_count < 4:
                    result.s_partial += 1
                if record.polarity == "正极":
                    result.positive_count += 1
                elif record.polarity == "负极":
                    result.negative_count += 1
                reason = f"试样={record.sample_name}, 极性={record.polarity}, 曲线{curve_count}条"
                result.history.append(FileHistory(
                    file_path=fpath,
                    file_name=os.path.basename(fpath),
                    success=True,
                    reason=reason,
                    operation_time=op_time,
                ))
                logger.debug(f"Excel 提取成功: {record.sample_name} 曲线数={curve_count}")
            else:
                result.fail_count += 1
                result.history.append(FileHistory(
                    file_path=fpath,
                    file_name=os.path.basename(fpath),
                    success=False,
                    reason="未提取到有效数据（缺少试样名称或剥离强度曲线）",
                    operation_time=op_time,
                ))
                logger.debug(f"Excel 提取失败: {os.path.basename(fpath)}")

        # 解析 CSV
        for fpath in csv_files:
            if self._cancelled:
                break
            processed += 1
            if progress_callback:
                progress_callback(
                    processed, result.total_files,
                    f"正在解析: {os.path.basename(fpath)}"
                )

            record = self._excel_parser.parse(fpath)
            if record:
                all_records.append(record)
                result.success_count += 1
                curve_count = sum(1 for i in range(1, 10) if getattr(record, f"curve_{i}") is not None)
                result.s_complete += 1 if curve_count >= 4 else (1 if curve_count > 0 else 0)
                if curve_count > 0 and curve_count < 4:
                    result.s_partial += 1
                if record.polarity == "正极":
                    result.positive_count += 1
                elif record.polarity == "负极":
                    result.negative_count += 1
                reason = f"试样={record.sample_name}, 极性={record.polarity}, 曲线{curve_count}条"
                result.history.append(FileHistory(
                    file_path=fpath,
                    file_name=os.path.basename(fpath),
                    success=True,
                    reason=reason,
                    operation_time=op_time,
                ))
                logger.debug(f"CSV 提取成功: {record.sample_name} 曲线数={curve_count}")
            else:
                result.fail_count += 1
                result.history.append(FileHistory(
                    file_path=fpath,
                    file_name=os.path.basename(fpath),
                    success=False,
                    reason="未提取到有效数据（缺少试样名称或剥离强度曲线）",
                    operation_time=op_time,
                ))
                logger.debug(f"CSV 提取失败: {os.path.basename(fpath)}")

        # 去重：同一 (test_date, test_time, polarity) 只保留一条
        # PDF 和 Excel 提取的 sample_name 可能不同，但代表同一物理测试
        all_records, skipped = self._dedup_records(all_records)
        result.records = all_records
        result.skipped_records = skipped
        result.app_skipped = len(skipped)
        # 同步统计计数：去重后 success_count 应为去重后的记录数
        dedup_removed = result.success_count - len(all_records)
        if dedup_removed > 0:
            result.success_count = len(all_records)

        # 写入数据库
        if save_to_db and all_records and not self._cancelled:
            if progress_callback:
                progress_callback(
                    result.total_files, result.total_files,
                    "正在写入数据库..."
                )
            result.db_inserted, result.db_skipped = self._save_to_db(all_records)

        if self._cancelled:
            logger.info("提取已被取消")

        # 将历史记录持久化到数据库
        self._save_history_to_db(result.history, request_id)

        logger.info(result.summary)
        return result

    def _dedup_records(
        self, records: List[PeelDataRecord]
    ) -> Tuple[List[PeelDataRecord], List[SkippedRecord]]:
        """
        去重：按 (test_date, test_time, polarity) 去重

        同一时间同一极性的物理测试，可能从不同文件（PDF/Excel/转换Excel）
        提取到不同的 sample_name，但它们代表同一次测试，只应保留一条。

        规则：
        1. 批次内去重：相同 key 只保留第一条，其余记入 skipped
        2. 跨批次去重已移除——由数据库 UNIQUE 约束在写入时自动处理

        Returns:
            (去重后保留的记录列表, 被跳过的记录列表（含原因）)
        """
        if not records:
            return records, []

        # 批次内去重：相同 (test_date, test_time, polarity) 只保留第一条
        seen: dict = {}     # key -> 保留的 record
        skipped: List[SkippedRecord] = []
        for r in records:
            d = r.to_dict()
            key = (d["test_date"], d["test_time"], d["polarity"])
            if key in seen:
                skipped.append(SkippedRecord(
                    record=r,
                    reason=f"批次内重复（{seen[key].sample_name} 已保留）",
                ))
            else:
                seen[key] = r

        internal_dedup = len(records) - len(seen)
        if internal_dedup > 0:
            logger.info(f"批次内去重: 跳过 {internal_dedup} 条 (test_date+test_time+polarity 相同)")

        return list(seen.values()), skipped

    def _save_history_to_db(self, history: List[FileHistory], request_id: str):
        """将提取历史写入 peel_data_extraction_history 表
        【v1.7.4 全局历史解耦】写入时带 `plugin` 字段（默认 peel_data），
        让全局历史页可按插件筛选；不破坏旧数据的展示。
        """
        if not history:
            return
        try:
            ensure_history_table()
            table_name = get_history_table_name()
            db = DatabaseManager()
            for h in history:
                db.execute(
                    f'INSERT OR IGNORE INTO "{table_name}" '
                    '("file_path", "file_name", "success", "reason", "request_id", "plugin", "operation_time") '
                    "VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (
                        h.file_path,
                        h.file_name,
                        1 if h.success else 0,
                        h.reason,
                        request_id,
                        # 显式写明插件名（peel_data），由调用方负责在后续多插件场景中替换
                        "peel_data",
                        h.operation_time,
                    ),
                )
            logger.info(f"历史记录已持久化: {len(history)} 条 (plugin=peel_data)")
        except Exception as e:
            logger.error(f"历史记录持久化失败: {e}", exc_info=True)

    def _save_to_db(
        self, records: List[PeelDataRecord]
    ) -> Tuple[int, int]:
        """将记录批量写入数据库；DB 不可用时返回 (0, len(records))"""
        if not _is_db_available():
            logger.warning("数据库不可用，跳过写入")
            return 0, len(records)

        db = DatabaseManager()
        data_list = [r.to_dict() for r in records]
        try:
            inserted, skipped = db.insert_many_ignore(
                PeelDataRecord.get_table_name(), data_list
            )
            logger.info(f"数据库写入: 插入 {inserted}, 跳过重复 {skipped}")
            return inserted, skipped
        except Exception as e:
            logger.error(f"数据库写入失败: {e}", exc_info=True)
            # 尝试逐条插入
            inserted, skipped = 0, 0
            for data in data_list:
                try:
                    if db.insert_ignore(PeelDataRecord.get_table_name(), data):
                        inserted += 1
                    else:
                        skipped += 1
                except Exception as e2:
                    logger.error(f"单条插入失败: {e2}")
                    skipped += 1
            return inserted, skipped

    def cancel(self):
        """取消当前提取操作"""
        self._cancelled = True
        logger.info("用户取消提取操作")

    @staticmethod
    def export_to_excel(
        records: List[PeelDataRecord],
        output_path: str,
        split_polarity: bool = False,
    ) -> bool:
        """
        导出数据到 Excel 文件

        Args:
            records: 记录列表
            output_path: 输出文件路径
            split_polarity: 是否按正负极分 Sheet（默认 False）

        Returns:
            bool: 是否成功
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error("openpyxl 未安装，无法导出")
            return False

        try:
            wb = openpyxl.Workbook()

            # 确定实际存在的曲线列
            active_curves = set()
            for r in records:
                for i in range(1, 10):
                    if getattr(r, f"curve_{i}") is not None:
                        active_curves.add(i)
            active_curves = sorted(active_curves)

            # 表头
            headers = [
                "试样名称", "试样牌号", "极性", "试验日期时间",
            ]
            # 获取最常见的单位用于表头
            from collections import Counter
            units = [r.curve_unit for r in records if r.curve_unit]
            unit_suffix = ""
            if units:
                common_unit = Counter(units).most_common(1)[0][0]
                unit_suffix = f" ({common_unit})"
            curve_headers = [f"曲线{i}{unit_suffix}" for i in active_curves]
            headers.extend(curve_headers)
            headers.extend(["标准差", "来源文件", "文件类型"])

            # 样式定义
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            if split_polarity:
                # 按正负极分 Sheet
                pos_records = [r for r in records if r.polarity == "正极"]
                neg_records = [r for r in records if r.polarity == "负极"]
                other_records = [r for r in records if r.polarity not in ("正极", "负极")]

                # 正极 Sheet
                if pos_records:
                    ws = wb.active
                    ws.title = "正极"
                    PeelDataExtractor._write_sheet(ws, pos_records, headers, active_curves, header_font_white, header_fill, center_align, thin_border)
                else:
                    wb.remove(wb.active)

                # 负极 Sheet
                if neg_records:
                    ws = wb.create_sheet("负极")
                    PeelDataExtractor._write_sheet(ws, neg_records, headers, active_curves, header_font_white, header_fill, center_align, thin_border)

                # 其他 Sheet（如果有）
                if other_records:
                    ws = wb.create_sheet("其他")
                    PeelDataExtractor._write_sheet(ws, other_records, headers, active_curves, header_font_white, header_fill, center_align, thin_border)

                if not pos_records and not neg_records and not other_records:
                    # 空数据，创建一个空 Sheet
                    ws = wb.active
                    ws.title = "剥离数据汇总"

            else:
                # 原逻辑：单个 Sheet
                ws = wb.active
                ws.title = "剥离数据汇总"
                PeelDataExtractor._write_sheet(ws, records, headers, active_curves, header_font_white, header_fill, center_align, thin_border)

            wb.save(output_path)
            logger.info(f"数据导出成功: {output_path}")
            return True

        except Exception as e:
            logger.error(f"导出失败: {e}", exc_info=True)
            return False

    @staticmethod
    def _write_sheet(ws, records, headers, active_curves, header_font, header_fill, center_align, thin_border):
        """向工作表写入数据（辅助方法）"""
        # 写表头
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 写数据行
        for row_idx, record in enumerate(records, 2):
            col = 1
            # 试样名称
            ws.cell(row=row_idx, column=col, value=record.sample_name).border = thin_border
            col += 1
            # 试样牌号
            ws.cell(row=row_idx, column=col, value=record.sample_brand).border = thin_border
            col += 1
            # 极性
            ws.cell(row=row_idx, column=col, value=record.polarity).border = thin_border
            col += 1
            # 试验日期时间（合并显示）
            ws.cell(row=row_idx, column=col, value=record.test_datetime).border = thin_border
            col += 1

            # 曲线值（仅输出实际存在的曲线列）
            for i in active_curves:
                val = getattr(record, f"curve_{i}", None)
                ws.cell(row=row_idx, column=col, value=val).border = thin_border
                col += 1

            # 标准差
            ws.cell(row=row_idx, column=col, value=record.std_dev).border = thin_border
            col += 1
            # 来源文件
            ws.cell(row=row_idx, column=col, value=record.source_file).border = thin_border
            col += 1
            # 文件类型
            ws.cell(row=row_idx, column=col, value=record.file_type).border = thin_border

        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 4, 30)

    @staticmethod
    def export_to_csv(
        records: List[PeelDataRecord],
        output_path: str,
    ) -> bool:
        """
        导出数据到 CSV 文件

        Args:
            records: 记录列表
            output_path: 输出文件路径

        Returns:
            bool: 是否成功
        """
        try:
            import csv
        except ImportError:
            logger.error("csv 模块不可用")
            return False

        try:
            # 确定实际存在的曲线列
            active_curves = set()
            for r in records:
                for i in range(1, 10):
                    if getattr(r, f"curve_{i}") is not None:
                        active_curves.add(i)
            active_curves = sorted(active_curves)

            # 表头
            headers = [
                "试样名称", "试样牌号", "极性", "试验日期时间",
            ]
            curve_headers = [f"曲线{i}" for i in active_curves]
            headers.extend(curve_headers)
            headers.extend(["标准差", "来源文件", "文件类型"])

            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(headers)

                for record in records:
                    row = [
                        record.sample_name,
                        record.sample_brand,
                        record.polarity,
                        record.test_datetime,
                    ]
                    for i in active_curves:
                        row.append(getattr(record, f"curve_{i}"))
                    row.append(record.std_dev)
                    row.append(record.source_file)
                    row.append(record.file_type)
                    writer.writerow(row)

            logger.info(f"CSV 导出成功: {output_path}")
            return True

        except Exception as e:
            logger.error(f"CSV 导出失败: {e}", exc_info=True)
            return False
