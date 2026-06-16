# -*- coding: utf-8 -*-
"""分容数据解析器（精捷能设备 Excel/CSV 格式）

【v1.8.0 工艺算法 v5 — 由 Z 在 2026-06-17 确认】
普通分容流程：充电 → 放电 → 充电
- 第 1 步：充至满电态（恒流恒压充电）
- 第 2 步：执行 1 次完整放电（恒流放电）— 此为分容工步
- 第 3 步：执行 1 次充电（可能达满电态）

【关键判定 — 不依赖工步号，按关键字匹配】
- 恒流放电：描述含 "恒流放电"
- 恒流充电：描述含 "恒流充电" / "恒流恒压充电"

【容量取数】
取最后一次"恒流放电"工步的"单步容量(mAh)"
"""

from __future__ import annotations

import os
import re
from typing import List, Dict, Tuple, Optional

from core.logger import get_logger
from plugins.capacity_analysis.models import (
    DISCHARGE_KEYWORDS,
    CHARGE_KEYWORDS,
    CapacityRecord,
    AnalysisResult,
)

logger = get_logger("capacity_analysis.parser")


# 表头识别：单步容量列
_COL_SINGLE_CAPACITY = "单步容量(mAh)"
_COL_CUMULATIVE_CAPACITY = "累计容量(mAh)"
_COL_CELL_BARCODE = "电池条码"
_COL_STEP_NO = "数据内容"

# 第 1 列"工步编号"列名（包含在 _COL_STEP_NO 行）
_STEP_ID_RE = re.compile(r"(\d+)\((\d+)\)(.+)")


def _classify_step(step_desc: str) -> str:
    """根据工步描述返回工步类型。返回 "discharge" / "charge" / "rest" / "other" """
    if not step_desc:
        return "other"
    desc = str(step_desc)
    for kw in DISCHARGE_KEYWORDS:
        if kw in desc:
            return "discharge"
    for kw in CHARGE_KEYWORDS:
        if kw in desc:
            return "charge"
    if "静置" in desc or "搁置" in desc or "延时" in desc:
        return "rest"
    return "other"


def _parse_step_info(step_text: str) -> Tuple[int, int, str]:
    """解析"01_01 (001)"或"3(1)恒流恒压充电..."格式
    返回 (file_step_id, cycle_no, description)
    """
    if not step_text:
        return 0, 0, ""
    s = str(step_text)
    m = _STEP_ID_RE.match(s)
    if m:
        return int(m.group(1)), int(m.group(2)), m.group(3).strip()
    # 工步描述格式
    m2 = re.match(r"(\d+)\((\d+)\)\s*(.+)", s)
    if m2:
        return int(m2.group(1)), int(m2.group(2)), m2.group(3).strip()
    return 0, 0, s


def parse_capacity_file(file_path: str) -> AnalysisResult:
    """解析单个分容数据文件，返回分析结果

    Args:
        file_path: Excel/CSV 文件绝对路径

    Returns:
        AnalysisResult: 含 records / abnormal / stats
    """
    file_name = os.path.basename(file_path)
    batch_id = os.path.splitext(file_name)[0]

    result = AnalysisResult(batch_id=batch_id, source_file=file_path)
    result.abnormal = {
        "no_discharge": [],
        "no_charge": [],
        "order_error": [],
        "cycle_test": [],
        "null_value": [],
    }

    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        rows = _read_excel(file_path)
    elif ext == ".csv":
        rows = _read_csv(file_path)
    else:
        logger.error(f"不支持的文件格式: {file_path}")
        return result

    if not rows:
        logger.error(f"文件为空或解析失败: {file_path}")
        return result

    # 解析表头
    if len(rows) < 3:
        logger.error(f"文件行数不足: {file_path}")
        return result
    # 精捷能格式：第 1 行 = 工作状态，第 2 行 = 表头（数据内容/电池条码/...）
    # _read_excel / _read_csv 已经把这两行都保留在 rows 里
    header = rows[1]
    try:
        col_step_id = header.index(_COL_STEP_NO)
        col_capacity = header.index(_COL_SINGLE_CAPACITY)
    except ValueError as e:
        logger.error(f"表头缺少必要列 {_COL_STEP_NO}/{_COL_SINGLE_CAPACITY}: {e}")
        return result

    # 第 1 行是"工作状态"行（精捷能格式），从中提取工步描述
    step_descs: List[Tuple[int, str]] = []  # (起始列索引, 工步描述)
    if len(rows) >= 1:
        status_row = rows[0]
        for ci in range(2, len(status_row)):
            v = status_row[ci]
            if v and isinstance(v, str) and _STEP_ID_RE.search(v):
                step_descs.append((ci, v))
            elif v and isinstance(v, str) and re.search(r"\d+\(\d+\)", v):
                step_descs.append((ci, v))
    if not step_descs:
        logger.error(f"未在第 1 行找到任何工步描述: {file_path}")
        return result

    # 解析表头中所有"单步容量(mAh)"列位置 — 每个工步块都有独立的"单步容量"列
    # 关键设计：直接按"单步容量列位置"反推该列属于哪个工步
    capacity_cols: List[int] = [i for i, v in enumerate(header) if v == _COL_SINGLE_CAPACITY]
    if not capacity_cols:
        logger.error(f"表头未找到 '{_COL_SINGLE_CAPACITY}' 列")
        return result

    def _col_to_step_type(col_idx: int) -> Optional[str]:
        """根据容量列在表头中的位置，反推该列所属工步的类型。
        精捷能 Excel 列布局：每个工步块有 21 列（"数据内容"列单独占第 0 列不重复）。
        工步位置在 status_row 中，列布局紧跟其后。
        关键规则：col_idx 之前最近的"工步"位置 = 该列所属工步。
        """
        prev_step = None
        for sc, sd in step_descs:
            if sc <= col_idx:
                prev_step = (sc, sd)
            else:
                break
        if prev_step is None:
            return None
        return _classify_step(prev_step[1])

    # 数据行：第 3 行起（前 2 行是状态/表头）
    for ri, row in enumerate(rows[2:], start=3):
        if not row or not row[col_step_id]:
            continue
        cell_id = str(row[col_step_id]).strip()
        if not cell_id:
            continue
        # 解析该行所有工步 — 按"单步容量列"扫描，按列位置反推工步类型
        discharge_caps = []  # (cycle_no, capacity, step_desc)
        charge_count = 0
        for cap_col in capacity_cols:
            if cap_col >= len(row):
                continue
            step_type = _col_to_step_type(cap_col)
            if step_type is None:
                continue
            if step_type == "discharge":
                try:
                    val = row[cap_col]
                    if val is None or val == "" or val == 0:
                        continue
                    cap = float(val)
                    if cap <= 0:
                        continue
                    # 从 step_descs 找最近的工步描述
                    prev_step = None
                    for sc, sd in step_descs:
                        if sc <= cap_col:
                            prev_step = (sc, sd)
                        else:
                            break
                    desc = prev_step[1] if prev_step else ""
                    _, cyc, _ = _parse_step_info(desc)
                    discharge_caps.append((cyc, cap, desc))
                except (ValueError, TypeError, IndexError):
                    continue
            elif step_type == "charge":
                charge_count += 1
        charge_found = charge_count > 0

        # 简化兜底：如果上面没找到放电容量，用"累计容量"列最末非零值
        if not discharge_caps and _COL_CUMULATIVE_CAPACITY in header:
            cum_positions = [i for i, v in enumerate(header) if v == _COL_CUMULATIVE_CAPACITY]
            for cum_col in cum_positions:
                if cum_col >= len(row):
                    continue
                step_type = _col_to_step_type(cum_col)
                if step_type == "discharge":
                    try:
                        val = row[cum_col]
                        if val:
                            cap = float(val)
                            if cap > 0:
                                prev_step = None
                                for sc, sd in step_descs:
                                    if sc <= cum_col:
                                        prev_step = (sc, sd)
                                    else:
                                        break
                                desc = prev_step[1] if prev_step else ""
                                discharge_caps.append((99, cap, desc))
                    except (ValueError, TypeError, IndexError):
                        continue

        # 判定异常
        if not discharge_caps:
            if not charge_found:
                result.abnormal["no_discharge"].append({"row": ri, "cell_id": cell_id})
            else:
                result.abnormal["no_discharge"].append({"row": ri, "cell_id": cell_id})
            continue
        if not charge_found:
            result.abnormal["no_charge"].append({"row": ri, "cell_id": cell_id})
            continue
        # 顺序判定：第一次充电 < 第一次放电（已在 step_descs 顺序中保证）
        # 多次分容：>= 2 次放电视为"多次分容"（正常工艺），不再剔除
        # —— 决策 2（按分容次数分组）：多次分容仍有效，cycle_count 字段记录次数
        if len(discharge_caps) >= 2:
            # 多次分容仍保留（Z 工艺：金标准 1 次，但实际可能 2-3 次，按次数分组展示）
            pass
        # 取最后一次放电容量
        last_cap = discharge_caps[-1][1]
        if last_cap is None or last_cap == 0:
            result.abnormal["null_value"].append({"row": ri, "cell_id": cell_id})
            continue
        result.records.append(CapacityRecord(
            cell_id=cell_id,
            capacity_mah=last_cap,
            cycle_count=len(discharge_caps),
            file_name=file_name,
            source_row=ri,
        ))

    return result


def _read_excel(file_path: str) -> List[List]:
    """读取 Excel 文件为二维列表（read_only 模式）"""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl 未安装")
        return []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        # 取第一个非空 sheet
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(v is not None and str(v).strip() for v in row):
                    rows.append(list(row))
            if rows:
                wb.close()
                return rows
        wb.close()
        return []
    except Exception as e:
        logger.error(f"读取 Excel 失败: {e}", exc_info=True)
        return []


def _read_csv(file_path: str) -> List[List]:
    """读取 CSV 文件为二维列表"""
    import csv
    rows = []
    try:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                if any(v is not None and str(v).strip() for v in row):
                    rows.append(row)
    except UnicodeDecodeError:
        try:
            with open(file_path, "r", encoding="gbk", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    if any(v is not None and str(v).strip() for v in row):
                        rows.append(row)
        except Exception as e:
            logger.error(f"读取 CSV 失败: {e}", exc_info=True)
    except Exception as e:
        logger.error(f"读取 CSV 失败: {e}", exc_info=True)
    return rows


def _read_step_descs(rows: List[List], header: List) -> List[Tuple[int, str]]:
    """从"工作状态"行（第 1 行）读取每个工步的列起始索引和工步描述

    精捷能格式：第 1 行是工作状态行，记录每个工步的列起始位置和描述
    简化策略：扫描所有列，找出非空且像"3(1)恒流充电..."的字串
    """
    if len(rows) < 2:
        return []
    status_row = rows[1]
    # 工作状态行第 1 列是"工作状态"标签，从第 2 列开始是工步
    step_descs = []
    for ci in range(2, len(status_row)):
        v = status_row[ci]
        if v and isinstance(v, str) and _STEP_ID_RE.search(v):
            step_descs.append((ci, v))
        elif v and isinstance(v, str) and re.search(r"\d+\(\d+\)", v):
            step_descs.append((ci, v))
    return step_descs
